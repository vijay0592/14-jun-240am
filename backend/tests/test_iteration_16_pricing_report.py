"""
Iteration 16 backend tests:
- Price-list CRUD (admin only) + per-item price + per-category discount
- Excel export / import
- Customer transport_name + price_list_id fields
- Dispatch execute pricing enrichment (unit_price, discount, net_unit_price)
- Daily dispatch report grouped by party
- Auth gating (non-admin cannot mutate price-lists)
"""
import io
import os
import uuid
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://build-from-git-2.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"


# -------------------- fixtures --------------------
@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{API}/auth/login", json={"email": "admin", "password": "admin123"}, timeout=30)
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def user_token():
    r = requests.post(f"{API}/auth/login", json={"email": "user", "password": "user123"}, timeout=30)
    assert r.status_code == 200, f"user login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def admin_h(admin_token):
    return {"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"}


@pytest.fixture(scope="session")
def user_h(user_token):
    return {"Authorization": f"Bearer {user_token}", "Content-Type": "application/json"}


@pytest.fixture(scope="session")
def some_items(admin_h):
    """Fetch a few existing SKUs to use for pricing tests."""
    r = requests.get(f"{API}/items", headers=admin_h, timeout=30)
    assert r.status_code == 200, f"items list failed: {r.status_code} {r.text}"
    items = r.json()
    assert len(items) >= 2, "need at least 2 SKUs seeded"
    return items[:5]


@pytest.fixture(scope="session")
def created_price_list(admin_h):
    """Create a fresh price list once for the whole test run."""
    name = f"TEST_PL_{uuid.uuid4().hex[:6]}"
    r = requests.post(f"{API}/price-lists", headers=admin_h, json={"name": name}, timeout=30)
    assert r.status_code in (200, 201), f"create price list failed: {r.status_code} {r.text}"
    pl = r.json()
    assert pl.get("id") and pl.get("name") == name
    yield pl
    # cleanup - unassign customers + delete
    requests.delete(f"{API}/price-lists/{pl['id']}", headers=admin_h, timeout=30)


# -------------------- Price-list CRUD --------------------
class TestPriceLists:
    def test_create_appears_in_list(self, admin_h, created_price_list):
        r = requests.get(f"{API}/price-lists", headers=admin_h, timeout=30)
        assert r.status_code == 200
        rows = r.json()
        match = [p for p in rows if p["id"] == created_price_list["id"]]
        assert match, "created price list missing in GET /api/price-lists"
        p = match[0]
        assert "items_count" in p and "discounts_count" in p
        assert isinstance(p["items_count"], int)
        assert isinstance(p["discounts_count"], int)

    def test_get_detail_lists_all_skus(self, admin_h, created_price_list):
        r = requests.get(f"{API}/price-lists/{created_price_list['id']}", headers=admin_h, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "price_list" in data and "items" in data and "discounts" in data
        assert isinstance(data["items"], list) and len(data["items"]) > 0
        sample = data["items"][0]
        for k in ("item_id", "item_name", "product_name", "price"):
            assert k in sample, f"missing {k} in item row"
        # default price for never-set items is 0
        unset = [i for i in data["items"] if i["price"] == 0]
        assert len(unset) > 0

    def test_set_item_price_and_verify_persistence(self, admin_h, created_price_list, some_items):
        item = some_items[0]
        r = requests.post(
            f"{API}/price-lists/{created_price_list['id']}/items",
            headers=admin_h,
            json={"item_id": item["id"], "price": 123.45},
            timeout=30,
        )
        assert r.status_code in (200, 201), r.text
        # re-fetch detail
        r2 = requests.get(f"{API}/price-lists/{created_price_list['id']}", headers=admin_h, timeout=30)
        assert r2.status_code == 200
        rows = r2.json()["items"]
        row = next((x for x in rows if x["item_id"] == item["id"]), None)
        assert row is not None
        assert float(row["price"]) == pytest.approx(123.45)
        # items_count >= 1
        rl = requests.get(f"{API}/price-lists", headers=admin_h, timeout=30).json()
        match = next(x for x in rl if x["id"] == created_price_list["id"])
        assert match["items_count"] >= 1

    def test_set_category_discount_rs(self, admin_h, created_price_list, some_items):
        cat = some_items[0].get("product_name") or "DefaultCategory"
        r = requests.post(
            f"{API}/price-lists/{created_price_list['id']}/discounts",
            headers=admin_h,
            json={"product_name": cat, "discount_value": 5, "discount_type": "₹"},
            timeout=30,
        )
        assert r.status_code in (200, 201), r.text
        d = requests.get(f"{API}/price-lists/{created_price_list['id']}", headers=admin_h, timeout=30).json()
        discs = d["discounts"]
        match = next((x for x in discs if x["product_name"] == cat), None)
        assert match is not None
        assert match["discount_type"] in ("₹", "%")
        assert float(match["discount_value"]) == 5.0

    def test_excel_export_xlsx(self, admin_h, created_price_list):
        r = requests.get(f"{API}/price-lists/{created_price_list['id']}/export", headers=admin_h, timeout=60)
        assert r.status_code == 200, r.text
        ctype = r.headers.get("content-type", "")
        assert "spreadsheet" in ctype or "openxml" in ctype, f"unexpected content-type {ctype}"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # first row should be a header
        first = list(next(ws.iter_rows(values_only=True)))
        assert first[0] and "item" in str(first[0]).lower()
        assert first[1] and "price" in str(first[1]).lower()

    def test_excel_import_known_and_unknown(self, admin_h, created_price_list, some_items):
        wb = Workbook()
        ws = wb.active
        ws.append(["Item Name", "Price"])
        # 2 known items
        ws.append([some_items[0]["name"], 77.0])
        ws.append([some_items[1]["name"], 88.0])
        # 1 unknown
        ws.append(["NONEXISTENT_ZXY_999", 50.0])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("upload.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        hdr = {"Authorization": admin_h["Authorization"]}
        r = requests.post(
            f"{API}/price-lists/{created_price_list['id']}/import",
            headers=hdr, files=files, timeout=60,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["updated"] == 2, body
        assert body["unknown_count"] == 1
        assert any("NONEXISTENT_ZXY_999" in u["item_name"] for u in body["unknown"])
        assert any("no matching SKU" in u["reason"] for u in body["unknown"])
        # verify prices landed
        d = requests.get(f"{API}/price-lists/{created_price_list['id']}", headers=admin_h, timeout=30).json()
        rows = d["items"]
        r0 = next(x for x in rows if x["item_id"] == some_items[0]["id"])
        r1 = next(x for x in rows if x["item_id"] == some_items[1]["id"])
        assert float(r0["price"]) == 77.0
        assert float(r1["price"]) == 88.0


# -------------------- Auth gating --------------------
class TestAuthGating:
    def test_user_can_view_price_lists(self, user_h):
        r = requests.get(f"{API}/price-lists", headers=user_h, timeout=30)
        assert r.status_code == 200

    def test_user_cannot_create_price_list(self, user_h):
        r = requests.post(f"{API}/price-lists", headers=user_h,
                          json={"name": f"TEST_DENIED_{uuid.uuid4().hex[:5]}"}, timeout=30)
        assert r.status_code == 403, f"expected 403, got {r.status_code}: {r.text}"

    def test_user_cannot_set_item_price(self, user_h, created_price_list, some_items):
        r = requests.post(
            f"{API}/price-lists/{created_price_list['id']}/items",
            headers=user_h,
            json={"item_id": some_items[0]["id"], "price": 1.0},
            timeout=30,
        )
        assert r.status_code == 403

    def test_user_cannot_delete_price_list(self, user_h, created_price_list):
        r = requests.delete(f"{API}/price-lists/{created_price_list['id']}", headers=user_h, timeout=30)
        assert r.status_code == 403

    def test_user_can_view_daily_report(self, user_h):
        r = requests.get(f"{API}/reports/daily-dispatch", headers=user_h, timeout=30)
        assert r.status_code == 200


# -------------------- Customer transport+price_list --------------------
class TestCustomerFields:
    @pytest.fixture(scope="class")
    def created_customer(self, admin_h, created_price_list):
        payload = {
            "name": f"TEST_CUST_{uuid.uuid4().hex[:6]}",
            "phone": "9999900000",
            "address": "Test addr",
            "transport_name": "ABC Roadlines",
            "price_list_id": created_price_list["id"],
        }
        r = requests.post(f"{API}/customers", headers=admin_h, json=payload, timeout=30)
        assert r.status_code in (200, 201), r.text
        cust = r.json()
        yield cust
        requests.delete(f"{API}/customers/{cust['id']}", headers=admin_h, timeout=30)

    def test_create_with_transport_and_price_list(self, admin_h, created_customer, created_price_list):
        assert created_customer.get("transport_name") == "ABC Roadlines"
        assert created_customer.get("price_list_id") == created_price_list["id"]
        # GET list
        rows = requests.get(f"{API}/customers", headers=admin_h, timeout=30).json()
        match = next((c for c in rows if c["id"] == created_customer["id"]), None)
        assert match is not None
        assert match["transport_name"] == "ABC Roadlines"
        assert match["price_list_id"] == created_price_list["id"]

    def test_patch_transport_only(self, admin_h, created_customer):
        r = requests.patch(f"{API}/customers/{created_customer['id']}", headers=admin_h,
                           json={"transport_name": "XYZ Logistics"}, timeout=30)
        assert r.status_code == 200, r.text
        got = requests.get(f"{API}/customers", headers=admin_h, timeout=30).json()
        match = next(c for c in got if c["id"] == created_customer["id"])
        assert match["transport_name"] == "XYZ Logistics"
        # price_list still set
        assert match["price_list_id"]

    def test_patch_price_list_only(self, admin_h, created_customer):
        # set price_list to None / clear is risky; let's just re-set to same
        r = requests.patch(f"{API}/customers/{created_customer['id']}", headers=admin_h,
                           json={"price_list_id": created_customer["price_list_id"]}, timeout=30)
        assert r.status_code == 200, r.text


# -------------------- Dispatch enrichment + Daily report --------------------
class TestDispatchPricingAndReport:
    @pytest.fixture(scope="class")
    def setup_world(self, admin_h, some_items, created_price_list):
        item = some_items[0]
        cat = item.get("product_name") or "DefaultCategory"
        # set price 100 and 10% discount
        requests.post(f"{API}/price-lists/{created_price_list['id']}/items",
                      headers=admin_h, json={"item_id": item["id"], "price": 100.0}, timeout=30)
        requests.post(f"{API}/price-lists/{created_price_list['id']}/discounts",
                      headers=admin_h,
                      json={"product_name": cat, "discount_value": 10, "discount_type": "%"}, timeout=30)
        # customer
        cust_payload = {
            "name": f"TEST_DISP_CUST_{uuid.uuid4().hex[:6]}",
            "phone": "8888800000",
            "address": "T",
            "transport_name": "TestTransport-Pvt",
            "price_list_id": created_price_list["id"],
        }
        cust = requests.post(f"{API}/customers", headers=admin_h, json=cust_payload, timeout=30).json()
        # create order
        order_payload = {
            "customer_id": cust["id"],
            "items": [{
                "item_id": item["id"],
                "item_name": item["name"],
                "product_name": item.get("product_name") or "",
                "variant": item.get("variant") or "",
                "quantity": 5,
            }],
        }
        order_resp = requests.post(f"{API}/orders", headers=admin_h, json=order_payload, timeout=30)
        assert order_resp.status_code in (200, 201), order_resp.text
        order = order_resp.json()
        yield {"customer": cust, "order": order, "item": item, "category": cat}
        # cleanup
        requests.delete(f"{API}/orders/{order['id']}", headers=admin_h, timeout=30)
        requests.delete(f"{API}/customers/{cust['id']}", headers=admin_h, timeout=30)

    def test_dispatch_execute_enriches_pricing(self, admin_h, setup_world):
        item = setup_world["item"]
        order = setup_world["order"]
        r = requests.post(
            f"{API}/dispatch/execute",
            headers=admin_h,
            json={"order_id": order["id"],
                  "allocations": [{"item_id": item["id"], "quantity": 2}]},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        d = body["dispatch"]
        assert d.get("transport_name") == "TestTransport-Pvt"
        assert "total_value" in d
        line = d["items"][0]
        assert float(line["unit_price"]) == 100.0
        assert float(line["discount_value"]) == 10.0
        assert line["discount_type"] == "%"
        # net = 100 * (1 - 10/100) = 90
        assert float(line["net_unit_price"]) == pytest.approx(90.0, rel=0.01)
        # total_value = 90 * 2 = 180
        assert float(d["total_value"]) == pytest.approx(180.0, rel=0.01)

    def test_daily_report_today(self, admin_h, setup_world):
        r = requests.get(f"{API}/reports/daily-dispatch", headers=admin_h, timeout=30)
        assert r.status_code == 200
        body = r.json()
        assert "groups" in body and "grand_total_value" in body
        cust_id = setup_world["customer"]["id"]
        grp = next((g for g in body["groups"] if g["customer_id"] == cust_id), None)
        assert grp is not None, "dispatched customer missing from today's report"
        assert grp["transport_name"] == "TestTransport-Pvt"
        assert grp["dispatch_count"] >= 1
        assert len(grp["lines"]) >= 1
        ln = grp["lines"][0]
        for k in ("unit_price", "discount_value", "net_unit_price", "line_value"):
            assert k in ln, f"missing {k} in report line"
        # grand_total_value == sum(line_value)
        total = round(sum(L["line_value"] for g in body["groups"] for L in g["lines"]), 2)
        assert float(body["grand_total_value"]) == pytest.approx(total, abs=0.05)

    def test_daily_report_past_date_empty(self, admin_h):
        r = requests.get(f"{API}/reports/daily-dispatch?date=2000-01-01", headers=admin_h, timeout=30)
        assert r.status_code == 200
        body = r.json()
        assert body["dispatch_count"] == 0
        assert float(body["grand_total_value"]) == 0.0
        assert body["groups"] == []

    def test_daily_report_invalid_date(self, admin_h):
        r = requests.get(f"{API}/reports/daily-dispatch?date=NOT-A-DATE", headers=admin_h, timeout=30)
        assert r.status_code == 400


# -------------------- Regression smoke --------------------
class TestRegression:
    def test_customers_list(self, admin_h):
        r = requests.get(f"{API}/customers", headers=admin_h, timeout=30)
        assert r.status_code == 200

    def test_orders_list(self, admin_h):
        r = requests.get(f"{API}/orders", headers=admin_h, timeout=30)
        assert r.status_code == 200

    def test_products_list(self, admin_h):
        r = requests.get(f"{API}/products", headers=admin_h, timeout=30)
        assert r.status_code == 200
