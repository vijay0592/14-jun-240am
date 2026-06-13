"""
Iteration 18 backend tests — Admin bulk customer mgmt + Admin Dispatch Ledger.

Covers:
- POST /api/customers/bulk-delete (empty / valid / blocker cases)
- GET /api/customers/import/template (xlsx)
- POST /api/customers/import (valid, dup-in-db, dup-in-file, missing header)
- GET /api/admin/dispatch-ledger (admin + non-admin)
- PATCH /api/dispatches/{id}/gr (admin + regular user + 404)
- GET /api/admin/dispatch-ledger/{id}/slip (admin + non-admin)
"""

import io
import os
import time
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
assert BASE_URL, "REACT_APP_BACKEND_URL must be set"

API = f"{BASE_URL}/api"


# ---------- fixtures ----------
@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{API}/auth/login", json={"email": "admin", "password": "admin123"}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="session")
def user_token():
    r = requests.post(f"{API}/auth/login", json={"email": "user", "password": "user123"}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture
def admin_h(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture
def user_h(user_token):
    return {"Authorization": f"Bearer {user_token}"}


def _mkxlsx(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _create_customer(admin_h, name, phone=""):
    r = requests.post(
        f"{API}/customers",
        json={"name": name, "phone": phone, "address": "", "city": "", "location": "", "transport_name": ""},
        headers=admin_h,
        timeout=15,
    )
    assert r.status_code in (200, 201), r.text
    return r.json()


def _delete_customer(admin_h, cid):
    requests.delete(f"{API}/customers/{cid}", headers=admin_h, timeout=15)


# ---------- bulk-delete ----------
class TestBulkDelete:
    def test_empty_ids_returns_400(self, admin_h):
        r = requests.post(f"{API}/customers/bulk-delete", json={"ids": []}, headers=admin_h, timeout=15)
        assert r.status_code == 400

    def test_bulk_delete_success(self, admin_h):
        ts = int(time.time())
        a = _create_customer(admin_h, f"TEST_BULK_A_{ts}")
        b = _create_customer(admin_h, f"TEST_BULK_B_{ts}")
        r = requests.post(
            f"{API}/customers/bulk-delete",
            json={"ids": [a["id"], b["id"]]},
            headers=admin_h,
            timeout=15,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("ok") is True
        assert data.get("deleted") == 2
        assert set(data.get("ids", [])) == {a["id"], b["id"]}
        # GET to verify removal
        r2 = requests.get(f"{API}/customers/search", params={"q": f"TEST_BULK_A_{ts}"}, headers=admin_h, timeout=15)
        names = [c.get("name") for c in (r2.json() if r2.status_code == 200 else [])]
        assert f"TEST_BULK_A_{ts}" not in names

    def test_bulk_delete_blocked_by_order(self, admin_h):
        # Find a customer that has an order (not off-order dispatch)
        # Use the dispatch ledger to find candidate customer ids and probe each.
        r = requests.get(f"{API}/admin/dispatch-ledger", params={"limit": 200}, headers=admin_h, timeout=20)
        cand_ids = []
        for it in (r.json().get("items", []) if r.status_code == 200 else []):
            cid = it.get("customer_id")
            if cid and cid not in cand_ids:
                cand_ids.append(cid)
        blocked_id = None
        for cid in cand_ids:
            # Probe by calling bulk-delete with just this id; if blocked → 400
            probe = requests.post(
                f"{API}/customers/bulk-delete", json={"ids": [cid]}, headers=admin_h, timeout=15
            )
            if probe.status_code == 400 and "Cannot delete" in (probe.text or ""):
                blocked_id = cid
                break
        if not blocked_id:
            pytest.skip("No customer with order references found in DB")
        ts = int(time.time())
        safe = _create_customer(admin_h, f"TEST_BULK_SAFE_{ts}")
        before = len(requests.get(f"{API}/customers", headers=admin_h, timeout=15).json())
        r = requests.post(
            f"{API}/customers/bulk-delete",
            json={"ids": [blocked_id, safe["id"]]},
            headers=admin_h,
            timeout=15,
        )
        assert r.status_code == 400, r.text
        after = len(requests.get(f"{API}/customers", headers=admin_h, timeout=15).json())
        assert after == before, "no customers should have been deleted"
        _delete_customer(admin_h, safe["id"])


# ---------- import template + import ----------
class TestCustomerImport:
    def test_template_download(self, admin_h):
        r = requests.get(f"{API}/customers/import/template", headers=admin_h, timeout=20)
        assert r.status_code == 200, r.text
        assert "openxmlformats" in r.headers.get("Content-Type", "")
        assert len(r.content) > 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for col in ["name", "phone", "address", "city", "location", "transport_name", "price_list"]:
            assert col in headers

    def test_import_2_new(self, admin_h):
        ts = int(time.time())
        n1, n2 = f"TEST_IMP_A_{ts}", f"TEST_IMP_B_{ts}"
        buf = _mkxlsx([
            ["name", "phone", "address", "city", "location", "transport_name", "price_list"],
            [n1, "9000000001", "addr1", "Indore", "loc1", "DTDC", ""],
            [n2, "9000000002", "addr2", "Bhopal", "loc2", "VRL", ""],
        ])
        r = requests.post(
            f"{API}/customers/import",
            files={"file": ("imp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_h,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        assert r.json().get("imported") == 2
        rg = requests.get(f"{API}/customers/search", params={"q": n1}, headers=admin_h, timeout=15)
        assert any(c["name"] == n1 for c in rg.json())
        # cleanup
        for c in requests.get(f"{API}/customers/search", params={"q": f"TEST_IMP_"}, headers=admin_h, timeout=15).json():
            _delete_customer(admin_h, c["id"])

    def test_import_duplicate_in_db(self, admin_h):
        ts = int(time.time())
        existing_name = f"TEST_DUPDB_{ts}"
        c = _create_customer(admin_h, existing_name)
        before = len(requests.get(f"{API}/customers", headers=admin_h, timeout=15).json())
        buf = _mkxlsx([
            ["name", "phone", "address", "city", "location", "transport_name", "price_list"],
            [existing_name, "9000000099", "x", "", "", "", ""],
            [f"TEST_DUPDB_NEW_{ts}", "9000000098", "x", "", "", "", ""],
        ])
        r = requests.post(
            f"{API}/customers/import",
            files={"file": ("imp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_h,
            timeout=30,
        )
        assert r.status_code == 400, r.text
        detail = r.json().get("detail")
        assert isinstance(detail, dict)
        assert "duplicates" in detail
        assert any(d.get("name") == existing_name for d in detail["duplicates"])
        after = len(requests.get(f"{API}/customers", headers=admin_h, timeout=15).json())
        assert after == before, "no customer should have been written"
        _delete_customer(admin_h, c["id"])

    def test_import_duplicate_in_file(self, admin_h):
        ts = int(time.time())
        name = f"TEST_DUPFILE_{ts}"
        buf = _mkxlsx([
            ["name", "phone", "address", "city", "location", "transport_name", "price_list"],
            [name, "", "", "", "", "", ""],
            [name, "", "", "", "", "", ""],
        ])
        r = requests.post(
            f"{API}/customers/import",
            files={"file": ("imp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_h,
            timeout=30,
        )
        assert r.status_code == 400, r.text
        detail = r.json().get("detail")
        assert isinstance(detail, dict)
        # row 3 = the duplicate of row 2
        rows = [d.get("row") for d in detail.get("duplicates", [])]
        assert "3" in rows

    def test_import_missing_name_header(self, admin_h):
        buf = _mkxlsx([
            ["phone", "address"],
            ["9999999999", "x"],
        ])
        r = requests.post(
            f"{API}/customers/import",
            files={"file": ("imp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_h,
            timeout=30,
        )
        assert r.status_code == 400, r.text


# ---------- dispatch ledger + GR + slip ----------
class TestDispatchLedger:
    def test_ledger_admin_ok(self, admin_h):
        r = requests.get(f"{API}/admin/dispatch-ledger", headers=admin_h, timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        for k in ("total", "items", "grand_total_value", "grand_total_pcs"):
            assert k in data
        # Filter params accepted
        r2 = requests.get(
            f"{API}/admin/dispatch-ledger",
            params={"start_date": "2026-01-01", "end_date": "2026-12-31", "limit": 5, "skip": 0},
            headers=admin_h,
            timeout=20,
        )
        assert r2.status_code == 200

    def test_ledger_user_forbidden(self, user_h):
        r = requests.get(f"{API}/admin/dispatch-ledger", headers=user_h, timeout=20)
        assert r.status_code == 403

    def test_patch_gr_admin_and_user(self, admin_h, user_h):
        # pick first dispatch
        r = requests.get(f"{API}/admin/dispatch-ledger", headers=admin_h, timeout=20)
        items = r.json().get("items", [])
        if not items:
            pytest.skip("No dispatch in DB to PATCH")
        did = items[0]["id"]
        ts = int(time.time())
        # admin sets
        r1 = requests.patch(f"{API}/dispatches/{did}/gr", json={"gr_number": f"ADM-{ts}"}, headers=admin_h, timeout=15)
        assert r1.status_code == 200, r1.text
        d1 = r1.json()
        assert d1.get("gr_number") == f"ADM-{ts}"
        assert d1.get("gr_updated_by")
        assert d1.get("gr_updated_at")
        # user updates
        r2 = requests.patch(f"{API}/dispatches/{did}/gr", json={"gr_number": f"USR-{ts}"}, headers=user_h, timeout=15)
        assert r2.status_code == 200, r2.text
        assert r2.json().get("gr_number") == f"USR-{ts}"

    def test_patch_gr_404(self, admin_h):
        r = requests.patch(f"{API}/dispatches/no-such-id/gr", json={"gr_number": "X"}, headers=admin_h, timeout=15)
        assert r.status_code == 404

    def test_slip_admin_and_user(self, admin_h, user_h):
        r = requests.get(f"{API}/admin/dispatch-ledger", headers=admin_h, timeout=20)
        items = r.json().get("items", [])
        if not items:
            pytest.skip("No dispatch in DB")
        did = items[0]["id"]
        ra = requests.get(f"{API}/admin/dispatch-ledger/{did}/slip", headers=admin_h, timeout=15)
        assert ra.status_code == 200, ra.text
        data = ra.json()
        assert "dispatch" in data and "customer" in data
        assert data["customer"].get("name")
        # Non-admin must be forbidden
        ru = requests.get(f"{API}/admin/dispatch-ledger/{did}/slip", headers=user_h, timeout=15)
        assert ru.status_code == 403
