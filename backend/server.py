from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Request, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import re
import base64
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from rapidfuzz import fuzz, process as rf_process, utils as rf_utils

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = os.environ.get('JWT_ALG', 'HS256')
JWT_EXPIRES_HOURS = int(os.environ.get('JWT_EXPIRES_HOURS', '24'))

app = FastAPI(title="Factory Order Management System")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ======================== Synonyms & Product Defaults ========================
SYNONYM_MAP = {
    # Center Stand With Kit
    "center stand with kit": "Center Stand with Kit",
    "main stand with kit": "Center Stand with Kit",
    "double stand with kit": "Center Stand with Kit",
    "kit": "Center Stand with Kit",
    # Center Stand Without Kit
    "center stand without kit": "Center Stand without Kit",
    "main stand without kit": "Center Stand without Kit",
    "double stand without kit": "Center Stand without Kit",
    "center stand": "Center Stand without Kit",
    "main stand": "Center Stand without Kit",
    # Center Stand Pin
    "center stand pin": "Center Stand Pin",
    "stand pin": "Center Stand Pin",
    "pin": "Center Stand Pin",
    # Seat Kunda
    "seat kunda": "Seat Kunda",
    "side seat handle": "Seat Kunda",
    # Lady footrest synonyms (user said lady footrest = side footrest but provided separate bag limits)
    "lady footrest": "Lady Footrest",
    # Side Stand
    "side stand": "Side Stand",
    "side stand splender": "Side Stand",
    # Other
    "footrest rod": "Footrest Rod",
    "front footrest rod": "Footrest Rod",
    "handlebar": "Handlebar",
    "handle bar": "Handlebar",
    "number plate": "Number Plate",
    "front number plate": "Number Plate",
    "rear number plate": "Number Plate",
    "no plate": "Number Plate",
    "engine plate": "Engine Plate",
    "side footrest": "Side Footrest",
    # New master products
    "v-bracket": "V-Bracket",
    "v bracket": "V-Bracket",
    "bracket": "V-Bracket",
    "luggage rod": "Luggage Rod",
    "luggage": "Luggage Rod",
    "side mirror clump": "Side Mirror Clump",
    "mirror clump": "Side Mirror Clump",
    "side mirror": "Side Mirror Clump",
    "rear seat handle": "Rear Seat Handle",
    "back seat handle": "Rear Seat Handle",
    "seat handle": "Rear Seat Handle",
    # Hindi / Devanagari product hints
    "साइड स्टैंड": "Side Stand",
    "सेंटर स्टैंड": "Center Stand without Kit",
    "सेंटर स्टैंड किट": "Center Stand with Kit",
    "मेन स्टैंड": "Center Stand without Kit",
    "पिन": "Center Stand Pin",
    "सीट कुंडा": "Seat Kunda",
    "फुटरेस्ट रॉड": "Footrest Rod",
    "लेडी फुटरेस्ट": "Lady Footrest",
    "हैंडलबार": "Handlebar",
    "हैंडल बार": "Handlebar",
    "नंबर प्लेट": "Number Plate",
    "इंजन प्लेट": "Engine Plate",
    "ब्रैकेट": "V-Bracket",
    "लगेज रॉड": "Luggage Rod",
    "मिरर": "Side Mirror Clump",
}

DEFAULT_PRODUCTS = [
    {"name": "Side Stand", "min_per_bag": 180, "max_per_bag": 200, "variants": ["Type A", "Type B", "Type C"], "variant_field": "side_stand_type"},
    {"name": "Center Stand with Kit", "min_per_bag": 55, "max_per_bag": 55, "variants": [], "variant_field": "center_stand_kit"},
    {"name": "Center Stand without Kit", "min_per_bag": 60, "max_per_bag": 60, "variants": [], "variant_field": "center_stand_kit"},
    {"name": "Center Stand Pin", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Footrest Rod", "min_per_bag": 70, "max_per_bag": 70, "variants": [], "variant_field": None},
    {"name": "Seat Kunda", "min_per_bag": 250, "max_per_bag": 300, "variants": ["Fix", "Folding"], "variant_field": "seat_kunda_type"},
    {"name": "Lady Footrest", "min_per_bag": 250, "max_per_bag": 300, "variants": [], "variant_field": None},
    {"name": "Handlebar", "min_per_bag": 90, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Number Plate", "min_per_bag": 300, "max_per_bag": 400, "variants": [], "variant_field": None},
    {"name": "Engine Plate", "min_per_bag": 200, "max_per_bag": 250, "variants": [], "variant_field": None},
    {"name": "Side Footrest", "min_per_bag": 50, "max_per_bag": 50, "variants": [], "variant_field": None},
    # New master products (Feb 2026) — bag limits default to 50/100, editable later
    {"name": "V-Bracket", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Luggage Rod", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Side Mirror Clump", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
    {"name": "Rear Seat Handle", "min_per_bag": 50, "max_per_bag": 100, "variants": [], "variant_field": None},
]

# Map item-sheet category headers → master product name in DB.
# User's explicit rules:
#  KIT → Center Stand with Kit, CENTER STAND (WITHOUT KIT) → Center Stand without Kit,
#  PIN → Center Stand Pin, FRONT/REAR NUMBER PLATE → Number Plate,
#  FRONT FOOTREST ROD → Footrest Rod, HANDLE BAR → Handlebar,
#  ENGINE PLATE → Engine Plate, V-BRACKET → V-Bracket,
#  LUGGAGE ROD → Luggage Rod, SIDE MIRROR CLUMP → Side Mirror Clump,
#  REAR SEAT HANDLE → Rear Seat Handle.
CATEGORY_TO_PRODUCT: Dict[str, str] = {
    "JK SIDE STAND": "Side Stand",
    "JK CENTER STAND": "Center Stand without Kit",
    "JK CENTER STAND KIT": "Center Stand with Kit",
    "JK CENTER STAND PIN": "Center Stand Pin",
    "JK FRONT NUMBER PLATE": "Number Plate",
    "JK REAR NUMBER PLATE": "Number Plate",
    "JK FRONT FOOTREST ROD": "Footrest Rod",
    "JK V-BRACKET": "V-Bracket",
    "JK HANDLE BAR": "Handlebar",
    "MOTER CYCLE ENGINE PLATE": "Engine Plate",
    "JK SEAT KUNDA": "Seat Kunda",
    "JK LADY FOOTREST": "Lady Footrest",
    "JK SIDE FOOTREST": "Side Footrest",
    "JK LUGGAGE ROD": "Luggage Rod",
    "JK SIDE MIRROR CLUMP": "Side Mirror Clump",
    "JK REAR SEAT HANDLE": "Rear Seat Handle",
}


# ======================== Models ========================
class UserIn(BaseModel):
    email: str  # accepts either email or username (validated downstream)
    password: str


class TokenOut(BaseModel):
    token: str
    user: Dict[str, Any]


class ProductIn(BaseModel):
    name: str
    min_per_bag: int
    max_per_bag: int
    variants: List[str] = []
    variant_field: Optional[str] = None


class ProductUpdate(BaseModel):
    min_per_bag: Optional[int] = None
    max_per_bag: Optional[int] = None
    variants: Optional[List[str]] = None


class ItemBagUpdate(BaseModel):
    # Per-SKU bag override. When set, overrides master product's bag limits
    # during dispatch bag calculation.
    min_per_bag: int
    max_per_bag: int


class CustomerIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    location: Optional[str] = ""
    preferences: Dict[str, str] = {}
    price_list_id: Optional[str] = None  # assigned price list (per-party pricing)
    transport_name: Optional[str] = ""   # transport company / vehicle name
    private_mark: Optional[str] = ""     # stenciled mark on packages for this party


class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    location: Optional[str] = None
    preferences: Optional[Dict[str, str]] = None
    price_list_id: Optional[str] = None
    transport_name: Optional[str] = None
    private_mark: Optional[str] = None


class OrderItemIn(BaseModel):
    product_name: str
    quantity: int
    variant: Optional[str] = None  # e.g. "Type A" for side stand
    # Strict item-wise: every order line MUST identify a specific SKU.
    item_id: str
    item_name: str


class OrderIn(BaseModel):
    customer_id: str
    items: List[OrderItemIn]
    order_date: Optional[str] = None
    delivery_date: Optional[str] = None
    notes: Optional[str] = ""
    merge_with_pending: bool = False
    clear_previous_pending: bool = False


class OrderStatusUpdate(BaseModel):
    status: str  # Pending / Dispatched / Cleared


class OrderUpdate(BaseModel):
    """Admin-only full-order edit. Any subset of fields can be supplied."""
    customer_id: Optional[str] = None
    items: Optional[List[OrderItemIn]] = None
    order_date: Optional[str] = None
    delivery_date: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None


class AdminUserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: str = "user"  # "admin" or "user"
    username: Optional[str] = None  # admin-supplied short login id (defaults to email local-part)


class AdminPasswordReset(BaseModel):
    password: str


class SettingsUpdate(BaseModel):
    overdue_days: int


class ItemCreate(BaseModel):
    name: str
    product_id: str
    min_per_bag: Optional[int] = None  # SKU bag override (optional)
    max_per_bag: Optional[int] = None


class ItemEdit(BaseModel):
    """Admin edit of a single item SKU: name / product mapping / bag override.
    Distinct from ItemBagUpdate which strictly enforces both bag fields."""
    name: Optional[str] = None
    product_id: Optional[str] = None
    min_per_bag: Optional[int] = None
    max_per_bag: Optional[int] = None


class DispatchStockIn(BaseModel):
    # Strict item-wise: keys are item_id (SKU), values are qty available
    items: Dict[str, int]


class DispatchAllocationIn(BaseModel):
    item_id: str
    quantity: int


class DispatchExecuteIn(BaseModel):
    """Partial / lot-wise dispatch: subtract the given quantities from one
    pending order. If everything in the order reaches 0, mark Dispatched.
    Otherwise the order stays Pending with reduced quantities so the
    remaining lot can be dispatched later."""
    order_id: str
    allocations: List[DispatchAllocationIn]
    notes: Optional[str] = ""


# ======================== Price List Models ========================
class PriceListIn(BaseModel):
    name: str
    description: Optional[str] = ""


class PriceListUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class PriceListItemIn(BaseModel):
    """Set/update the price for one item inside a price list."""
    item_id: str
    price: float


class CategoryDiscountIn(BaseModel):
    """Per-category (master product) discount within a price list.
    `discount_type` is either '₹' (flat rupees off) or '%' (percentage off)."""
    product_name: str
    discount_value: float
    discount_type: str  # '₹' or '%'


# ---- Login attestation (consent-based security capture) ----
class LoginAttestationIn(BaseModel):
    """Captured at login with the user's explicit consent.

    All fields are optional — the client posts whatever the user granted.
    If the user clicked Skip or denied permissions, set `consent=False`
    and the relevant `*_skipped` flag, leaving photo/location empty.
    """
    consent: bool = False
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy_meters: Optional[float] = None
    photo_b64: Optional[str] = None  # JPEG data URL or raw base64, capped server-side
    photo_skipped: bool = False
    location_skipped: bool = False
    error: Optional[str] = None  # free-text reason if permission was denied


# ---- Off-order (direct) dispatch ----
class OffOrderDispatchItemIn(BaseModel):
    item_id: str
    quantity: int


class OffOrderDispatchIn(BaseModel):
    """Dispatch a list of SKUs to a party that has no pending order.

    Either `customer_id` (existing party) OR a non-empty `customer_name`
    (walk-in / one-off) is required. `transport_name` overrides the
    customer's default; price list is taken from the customer if assigned.
    """
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    transport_name: Optional[str] = None
    items: List[OffOrderDispatchItemIn]
    notes: Optional[str] = None


# ---- Bulk customer admin ----
class CustomerBulkDeleteIn(BaseModel):
    """Admin: delete many customers in one request. Each id is validated
    and the call is rejected (with details) if any of them are referenced
    by an order."""
    ids: List[str]


# ---- Dispatch ledger (GR number edit) ----
class DispatchGrUpdate(BaseModel):
    gr_number: str


# ======================== Helpers ========================
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_token(user: Dict[str, Any]) -> str:
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRES_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Dict[str, Any]:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALG])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_admin(user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


def normalize_product_name(raw: str) -> Optional[str]:
    key = (raw or "").strip().lower()
    if not key:
        return None
    if key in SYNONYM_MAP:
        return SYNONYM_MAP[key]
    # fuzzy against synonyms
    match = rf_process.extractOne(key, list(SYNONYM_MAP.keys()), scorer=fuzz.WRatio)
    if match and match[1] >= 80:
        return SYNONYM_MAP[match[0]]
    # fuzzy against canonical product names
    canonical = list({v for v in SYNONYM_MAP.values()})
    match2 = rf_process.extractOne(key, canonical, scorer=fuzz.WRatio)
    if match2 and match2[1] >= 75:
        return match2[0]
    return None


# ======================== Seeding ========================
async def seed_db():
    # Users
    if await db.users.count_documents({}) == 0:
        admin = {"id": str(uuid.uuid4()), "email": "admin@factory.com", "username": "admin", "name": "Admin",
                 "password": hash_password("admin123"), "role": "admin", "created_at": now_iso()}
        user = {"id": str(uuid.uuid4()), "email": "user@factory.com", "username": "user", "name": "Operator",
                "password": hash_password("user123"), "role": "user", "created_at": now_iso()}
        await db.users.insert_many([admin, user])
        logger.info("Seeded default users")
    else:
        # Backfill username for any pre-existing user (local-part of email, deduped)
        seen = set(u.get("username") for u in await db.users.find({"username": {"$exists": True}}, {"_id": 0, "username": 1}).to_list(1000) if u.get("username"))
        async for u in db.users.find({"username": {"$exists": False}}, {"_id": 0}):
            base = (u.get("email", "") or "").split("@")[0].lower() or u["id"][:8]
            uname = base
            i = 1
            while uname in seen:
                i += 1
                uname = f"{base}{i}"
            seen.add(uname)
            await db.users.update_one({"id": u["id"]}, {"$set": {"username": uname}})
    # Settings singleton — overdue order threshold (admin-configurable)
    if await db.settings.count_documents({"id": "global"}) == 0:
        await db.settings.insert_one({"id": "global", "overdue_days": 15, "updated_at": now_iso()})
        logger.info("Seeded default settings (overdue_days=15)")
    # Products — additive: insert any from DEFAULT_PRODUCTS that don't exist yet
    existing_names = {p["name"] for p in await db.products.find({}, {"_id": 0, "name": 1}).to_list(1000)}
    new_products = [p for p in DEFAULT_PRODUCTS if p["name"] not in existing_names]
    if new_products:
        docs = [{"id": str(uuid.uuid4()), **p, "created_at": now_iso()} for p in new_products]
        await db.products.insert_many(docs)
        logger.info("Seeded %d new products", len(new_products))
    # Items — seed from data/items_parsed.json on first run
    if await db.items.count_documents({}) == 0:
        items_path = ROOT_DIR / "data" / "items_parsed.json"
        if items_path.exists():
            import json
            data = json.loads(items_path.read_text())
            # Build product name → id map
            prod_map = {p["name"]: p["id"] for p in await db.products.find({}, {"_id": 0}).to_list(1000)}
            docs = []
            unknown = set()
            for cat, items in data.items():
                product_name = CATEGORY_TO_PRODUCT.get(cat)
                if not product_name:
                    unknown.add(cat)
                    continue
                pid = prod_map.get(product_name)
                if not pid:
                    unknown.add(f"product missing: {product_name}")
                    continue
                for name in items:
                    docs.append({
                        "id": str(uuid.uuid4()),
                        "name": name,
                        "category": cat,
                        "product_id": pid,
                        "product_name": product_name,
                        "created_at": now_iso(),
                    })
            if docs:
                await db.items.insert_many(docs)
                logger.info("Seeded %d item SKUs", len(docs))
            if unknown:
                logger.warning("Item seed: unmapped categories: %s", unknown)


# ======================== Auth Routes ========================
@api_router.post("/auth/login", response_model=TokenOut)
async def login(body: UserIn):
    # Allow login by either email OR username (case-insensitive)
    ident = (body.email or "").strip().lower()
    if not ident:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    user = await db.users.find_one({"$or": [{"email": ident}, {"username": ident}]}, {"_id": 0})
    if not user or not verify_password(body.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(user)
    return {"token": token, "user": {"id": user["id"], "email": user["email"], "username": user.get("username"), "name": user.get("name", ""), "role": user["role"]}}


@api_router.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


# ======================== Admin: User Management ========================
@api_router.get("/users")
async def list_users(admin=Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).sort("created_at", 1).to_list(500)
    return users


@api_router.post("/users")
async def create_user(body: AdminUserCreate, admin=Depends(require_admin)):
    if body.role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="role must be 'admin' or 'user'")
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="A user with this email already exists")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    username = (body.username or "").strip().lower() or email.split("@")[0]
    # ensure username uniqueness
    base = username
    i = 1
    while await db.users.find_one({"username": username}):
        i += 1
        username = f"{base}{i}"
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "username": username,
        "name": body.name.strip() or email,
        "password": hash_password(body.password),
        "role": body.role,
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    return {"id": doc["id"], "email": doc["email"], "username": doc["username"], "name": doc["name"], "role": doc["role"], "created_at": doc["created_at"]}


@api_router.delete("/users/{uid}")
async def delete_user(uid: str, admin=Depends(require_admin)):
    if uid == admin["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    res = await db.users.delete_one({"id": uid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True, "deleted": uid}


@api_router.post("/users/{uid}/reset-password")
async def reset_user_password(uid: str, body: AdminPasswordReset, admin=Depends(require_admin)):
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    res = await db.users.update_one(
        {"id": uid},
        {"$set": {"password": hash_password(body.password), "updated_at": now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


# ======================== Admin: Settings (overdue threshold) ========================
async def _get_settings_doc() -> Dict[str, Any]:
    doc = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if not doc:
        doc = {"id": "global", "overdue_days": 15}
        await db.settings.insert_one({**doc, "updated_at": now_iso()})
    return doc


@api_router.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    return await _get_settings_doc()


@api_router.patch("/settings")
async def update_settings(body: SettingsUpdate, admin=Depends(require_admin)):
    if body.overdue_days <= 0 or body.overdue_days > 365:
        raise HTTPException(status_code=400, detail="overdue_days must be between 1 and 365")
    await db.settings.update_one(
        {"id": "global"},
        {"$set": {"overdue_days": body.overdue_days, "updated_at": now_iso()}},
        upsert=True,
    )
    return await _get_settings_doc()


# ======================== Products ========================
@api_router.get("/products")
async def list_products(user=Depends(get_current_user)):
    items = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return items


@api_router.post("/products")
async def create_product(body: ProductIn, user=Depends(require_admin)):
    if await db.products.find_one({"name": body.name}):
        raise HTTPException(status_code=400, detail="Product already exists")
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_iso()}
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/products/{pid}")
async def update_product(pid: str, body: ProductUpdate, user=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    # Server-side bag-limit sanity (client also validates)
    mn = update.get("min_per_bag")
    mx = update.get("max_per_bag")
    if mn is not None and mn <= 0:
        raise HTTPException(status_code=400, detail="min_per_bag must be > 0")
    if mx is not None and mx <= 0:
        raise HTTPException(status_code=400, detail="max_per_bag must be > 0")
    if mn is not None and mx is not None and mn > mx:
        raise HTTPException(status_code=400, detail="min_per_bag cannot exceed max_per_bag")
    res = await db.products.update_one({"id": pid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return await db.products.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_admin)):
    prod = await db.products.find_one({"id": pid})
    if not prod:
        raise HTTPException(status_code=404, detail="Product not found")
    # Block deletion if any item SKU or order still references it
    item_refs = await db.items.count_documents({"product_id": pid})
    if item_refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {item_refs} item SKU(s) still mapped to this product")
    order_refs = await db.orders.count_documents({"items.product_id": pid})
    if order_refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {order_refs} order(s) still reference this product")
    await db.products.delete_one({"id": pid})
    return {"ok": True, "deleted": pid}


# ======================== Item SKUs ========================
@api_router.get("/items")
async def list_items(product_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if product_id:
        q["product_id"] = product_id
    items = await db.items.find(q, {"_id": 0}).sort("name", 1).to_list(5000)
    return items


@api_router.get("/items/search")
async def search_items(q: str = "", product_id: Optional[str] = None, limit: int = 15,
                       user=Depends(get_current_user)):
    q = q.strip()
    filt = {}
    if product_id:
        filt["product_id"] = product_id
    all_items = await db.items.find(filt, {"_id": 0}).to_list(5000)
    if not q:
        return all_items[:limit]
    names = [it["name"] for it in all_items]
    matches = rf_process.extract(q, names, scorer=fuzz.WRatio,
                                  processor=rf_utils.default_process, limit=limit)
    out = []
    seen = set()
    short_q = len(q) < 5
    for name, score, idx in matches:
        if score < (35 if short_q else 45):
            continue
        it = all_items[idx]
        if it["id"] in seen:
            continue
        seen.add(it["id"])
        out.append({**it, "match_score": score})
    return out


@api_router.get("/items/{iid}")
async def get_item(iid: str, user=Depends(get_current_user)):
    it = await db.items.find_one({"id": iid}, {"_id": 0})
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    return it


@api_router.post("/items")
async def create_item(body: ItemCreate, admin=Depends(require_admin)):
    """Admin: create a new SKU under an existing master product."""
    prod = await db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not prod:
        raise HTTPException(status_code=404, detail="Product not found for product_id")
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Item name required")
    if await db.items.find_one({"name": name, "product_id": body.product_id}):
        raise HTTPException(status_code=400, detail="An item with this name already exists under this product")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "category": prod["name"],
        "product_id": body.product_id,
        "product_name": prod["name"],
        "created_at": now_iso(),
    }
    if body.min_per_bag is not None or body.max_per_bag is not None:
        mn = body.min_per_bag if body.min_per_bag is not None else body.max_per_bag
        mx = body.max_per_bag if body.max_per_bag is not None else body.min_per_bag
        if mn is None or mx is None or mn <= 0 or mx <= 0 or mn > mx:
            raise HTTPException(status_code=400, detail="Invalid bag override")
        doc["min_per_bag"] = mn
        doc["max_per_bag"] = mx
    await db.items.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/items/{iid}")
async def update_item(iid: str, body: ItemEdit, user=Depends(require_admin)):
    """Admin: update an item SKU. Accepts any subset of name / product_id /
    bag-override fields. Preserves prior behaviour: passing only min_per_bag +
    max_per_bag still sets the bag override unchanged."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    update: Dict[str, Any] = {}
    if body.name is not None:
        nm = body.name.strip()
        if not nm:
            raise HTTPException(status_code=400, detail="Item name cannot be empty")
        # uniqueness within same product
        dup = await db.items.find_one({"name": nm, "product_id": body.product_id or existing["product_id"], "id": {"$ne": iid}})
        if dup:
            raise HTTPException(status_code=400, detail="Another item with this name exists under the same product")
        update["name"] = nm
    if body.product_id is not None and body.product_id != existing.get("product_id"):
        prod = await db.products.find_one({"id": body.product_id}, {"_id": 0})
        if not prod:
            raise HTTPException(status_code=404, detail="Product not found for product_id")
        update["product_id"] = body.product_id
        update["product_name"] = prod["name"]
        update["category"] = prod["name"]
    if body.min_per_bag is not None and body.max_per_bag is not None:
        if body.min_per_bag <= 0 or body.max_per_bag <= 0:
            raise HTTPException(status_code=400, detail="Bag values must be > 0")
        if body.min_per_bag > body.max_per_bag:
            raise HTTPException(status_code=400, detail="min_per_bag cannot exceed max_per_bag")
        update["min_per_bag"] = body.min_per_bag
        update["max_per_bag"] = body.max_per_bag
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.items.update_one({"id": iid}, {"$set": update})
    return {"ok": True}


@api_router.delete("/items/{iid}")
async def delete_item(iid: str, admin=Depends(require_admin)):
    """Admin: delete an item SKU. Blocked if any order line references it."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    refs = await db.orders.count_documents({"items.item_id": iid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} order(s) reference this SKU")
    await db.items.delete_one({"id": iid})
    return {"ok": True, "deleted": iid}


@api_router.delete("/items/{iid}/bag-override")
async def clear_item_bag_override(iid: str, user=Depends(require_admin)):
    """Remove the per-SKU bag override so the item falls back to its master
    product's bag limits."""
    existing = await db.items.find_one({"id": iid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")
    await db.items.update_one(
        {"id": iid},
        {"$unset": {"min_per_bag": "", "max_per_bag": ""}},
    )
    return {"ok": True}


# ======================== Customers ========================
@api_router.get("/customers")
async def list_customers(user=Depends(get_current_user)):
    items = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api_router.get("/customers/search")
async def search_customers(q: str = "", user=Depends(get_current_user)):
    q = q.strip()
    if not q:
        return []
    all_c = await db.customers.find({}, {"_id": 0}).to_list(5000)
    names = [c["name"] for c in all_c]
    matches = rf_process.extract(q, names, scorer=fuzz.WRatio, processor=rf_utils.default_process, limit=8)
    out = []
    seen = set()
    short_q = len(q) < 6
    for name, score, idx in matches:
        if score < (40 if short_q else 50):
            continue
        cust = all_c[idx]
        if cust["id"] in seen:
            continue
        seen.add(cust["id"])
        out.append({**cust, "match_score": score})
    return out


@api_router.post("/customers")
async def create_customer(body: CustomerIn, user=Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_iso()}
    await db.customers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/customers/{cid}")
async def update_customer(cid: str, body: CustomerUpdate, user=Depends(get_current_user)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    # Users (non-admin) can only update preferences
    if user["role"] != "admin":
        update = {k: v for k, v in update.items() if k == "preferences"}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    res = await db.customers.update_one({"id": cid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return await db.customers.find_one({"id": cid}, {"_id": 0})


@api_router.delete("/customers/{cid}")
async def delete_customer(cid: str, admin=Depends(require_admin)):
    """Admin: delete a customer. Blocked if any order references this party."""
    existing = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")
    refs = await db.orders.count_documents({"customer_id": cid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} order(s) reference this customer")
    await db.customers.delete_one({"id": cid})
    return {"ok": True, "deleted": cid}


@api_router.post("/customers/bulk-delete")
async def bulk_delete_customers(body: CustomerBulkDeleteIn, admin=Depends(require_admin)):
    """Admin: delete many customers in a single call. The whole call is
    rejected if any of the supplied ids are referenced by an order — the
    response lists the blocking parties so the operator can review."""
    ids = [i for i in (body.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail="No customer ids supplied")
    # Validate every id exists
    existing = await db.customers.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(len(ids))
    existing_ids = {c["id"] for c in existing}
    missing = [i for i in ids if i not in existing_ids]
    if missing:
        raise HTTPException(status_code=404, detail=f"{len(missing)} customer(s) not found")
    # Find any blocking references in orders
    blockers = await db.orders.aggregate([
        {"$match": {"customer_id": {"$in": ids}}},
        {"$group": {"_id": "$customer_id", "count": {"$sum": 1}}},
    ]).to_list(len(ids))
    if blockers:
        name_map = {c["id"]: c["name"] for c in existing}
        details = ", ".join(f"{name_map.get(b['_id'], b['_id'])} ({b['count']})" for b in blockers)
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete: {len(blockers)} customer(s) have orders — {details}",
        )
    res = await db.customers.delete_many({"id": {"$in": ids}})
    return {"ok": True, "deleted": res.deleted_count, "ids": ids}


@api_router.get("/customers/import/template")
async def customer_import_template(admin=Depends(require_admin)):
    """Download a blank Excel template for bulk customer import."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["name", "phone", "address", "city", "location", "transport_name", "price_list"]
    ws.append(headers)
    # one example row to make the format obvious (will be skipped if empty)
    ws.append(["EXAMPLE PARTY PVT LTD", "9876543210", "12, Industrial Area", "Indore", "Sapna Sangeeta", "DTDC", ""])
    widths = [40, 16, 36, 18, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customers_import_template.xlsx"'},
    )


@api_router.post("/customers/import")
async def import_customers(file: UploadFile = File(...), admin=Depends(require_admin)):
    """Admin: bulk import customers from an Excel file.

    Columns recognised (case-insensitive header row required):
      name, phone, address, city, location, transport_name, price_list.

    Rules:
    - On ANY duplicate (same name case-insensitive, OR same non-empty phone)
      matching either an existing customer in the DB or another row in the
      same file, the whole import is rejected with a 400 listing the
      duplicates. Nothing is written.
    - `price_list` is matched by name (case-insensitive) against existing
      price lists; unknown names are reported as errors.
    - Empty rows are skipped.
    """
    from openpyxl import load_workbook
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty sheet")
    # Parse header
    header_row = rows[0]
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace(" ", "_")
        if key:
            header_map[key] = idx
    if "name" not in header_map:
        raise HTTPException(status_code=400, detail='Excel must have a "name" column in the first row')

    def cell(r, key: str) -> str:
        idx = header_map.get(key)
        if idx is None or idx >= len(r):
            return ""
        v = r[idx]
        if v is None:
            return ""
        return str(v).strip()

    # Pre-load existing customers and price lists for duplicate detection
    existing = await db.customers.find({}, {"_id": 0, "id": 1, "name": 1, "phone": 1}).to_list(20000)
    existing_names_lower = {(c.get("name") or "").strip().lower(): c.get("name") for c in existing if c.get("name")}
    existing_phones = {(c.get("phone") or "").strip(): c.get("name") for c in existing if (c.get("phone") or "").strip()}
    price_lists = await db.price_lists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    pl_by_name = {pl["name"].strip().lower(): pl["id"] for pl in price_lists}

    # Validate every data row
    duplicates: List[Dict[str, str]] = []
    unknown_pl: List[Dict[str, str]] = []
    parsed_rows: List[Dict[str, Any]] = []
    seen_name_in_file: Dict[str, int] = {}
    seen_phone_in_file: Dict[str, int] = {}

    for row_idx, r in enumerate(rows[1:], start=2):  # excel row #
        if not r or all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        name = cell(r, "name")
        if not name:
            continue
        name_l = name.lower()
        phone = cell(r, "phone")
        if name_l in existing_names_lower:
            duplicates.append({"row": str(row_idx), "name": name, "reason": "name already exists"})
            continue
        if name_l in seen_name_in_file:
            duplicates.append({"row": str(row_idx), "name": name, "reason": f"duplicate of row {seen_name_in_file[name_l]}"})
            continue
        if phone and phone in existing_phones:
            duplicates.append({"row": str(row_idx), "name": name, "reason": f"phone {phone} already used by {existing_phones[phone]}"})
            continue
        if phone and phone in seen_phone_in_file:
            duplicates.append({"row": str(row_idx), "name": name, "reason": f"phone duplicates row {seen_phone_in_file[phone]}"})
            continue

        pl_name = cell(r, "price_list")
        pl_id: Optional[str] = None
        if pl_name:
            pl_id = pl_by_name.get(pl_name.lower())
            if not pl_id:
                unknown_pl.append({"row": str(row_idx), "name": name, "price_list": pl_name})
                continue

        seen_name_in_file[name_l] = row_idx
        if phone:
            seen_phone_in_file[phone] = row_idx
        parsed_rows.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "phone": phone,
            "address": cell(r, "address"),
            "city": cell(r, "city"),
            "location": cell(r, "location"),
            "transport_name": cell(r, "transport_name"),
            "price_list_id": pl_id,
            "preferences": {},
            "created_at": now_iso(),
        })

    if duplicates:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"Import stopped: {len(duplicates)} duplicate row(s)",
                "duplicates": duplicates[:100],
            },
        )
    if unknown_pl:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"Import stopped: {len(unknown_pl)} row(s) reference an unknown price list",
                "unknown_price_lists": unknown_pl[:100],
            },
        )
    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No customer rows found in the file")

    await db.customers.insert_many(parsed_rows)
    return {"imported": len(parsed_rows)}


# ======================== Orders ========================
async def _persist_customer_prefs(customer_id: str, items: List[OrderItemIn]):
    """Memorize variant choices on customer for future orders."""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    pmap = {p["name"]: p for p in products}
    prefs_update = {}
    for it in items:
        prod = pmap.get(it.product_name)
        if not prod:
            continue
        vf = prod.get("variant_field")
        if not vf:
            continue
        if it.product_name in ("Center Stand with Kit", "Center Stand without Kit"):
            prefs_update["center_stand_kit"] = "With Kit" if "with Kit" in it.product_name else "Without Kit"
        elif it.variant:
            prefs_update[vf] = it.variant
    if prefs_update:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if cust:
            merged = {**(cust.get("preferences") or {}), **prefs_update}
            await db.customers.update_one({"id": customer_id}, {"$set": {"preferences": merged}})


@api_router.post("/orders")
async def create_order(body: OrderIn, user=Depends(get_current_user)):
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Strict item-wise validation: every line must have a valid item_id (SKU)
    if not body.items:
        raise HTTPException(status_code=400, detail="Order must contain at least one item")
    item_ids = [it.item_id for it in body.items]
    found = await db.items.find({"id": {"$in": item_ids}}, {"_id": 0, "id": 1}).to_list(1000)
    found_ids = {f["id"] for f in found}
    missing = [iid for iid in item_ids if iid not in found_ids]
    if missing:
        raise HTTPException(status_code=400, detail=f"Unknown item_id(s): {missing}")

    # If clear_previous_pending: mark all this customer's pending orders as Cleared
    if body.clear_previous_pending:
        await db.orders.update_many(
            {"customer_id": body.customer_id, "status": "Pending"},
            {"$set": {"status": "Cleared", "updated_at": now_iso()}},
        )

    # If merge_with_pending: append items to most recent pending order
    if body.merge_with_pending:
        existing = await db.orders.find_one(
            {"customer_id": body.customer_id, "status": "Pending"},
            sort=[("created_at", -1)],
        )
        if existing:
            new_items = existing.get("items", []) + [it.model_dump() for it in body.items]
            await db.orders.update_one(
                {"id": existing["id"]},
                {"$set": {"items": new_items, "updated_at": now_iso(),
                          "delivery_date": body.delivery_date or existing.get("delivery_date"),
                          "notes": (existing.get("notes", "") + " | " + (body.notes or "")).strip(" |")}},
            )
            await _persist_customer_prefs(body.customer_id, body.items)
            return await db.orders.find_one({"id": existing["id"]}, {"_id": 0})

    doc = {
        "id": str(uuid.uuid4()),
        "customer_id": body.customer_id,
        "customer_name": cust["name"],
        "items": [it.model_dump() for it in body.items],
        "order_date": body.order_date or now_iso(),
        "delivery_date": body.delivery_date,
        "status": "Pending",
        "notes": body.notes or "",
        "created_by": user["email"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.orders.insert_one(doc)
    await _persist_customer_prefs(body.customer_id, body.items)
    doc.pop("_id", None)
    return doc


@api_router.get("/orders")
async def list_orders(status_filter: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if status_filter:
        q["status"] = status_filter
    items = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    # Annotate overdue flag + days_open for Pending orders, using admin-set threshold
    settings = await _get_settings_doc()
    threshold = int(settings.get("overdue_days", 15))
    now = datetime.now(timezone.utc)
    for o in items:
        days_open = None
        ref = o.get("order_date") or o.get("created_at")
        if ref:
            try:
                dt = datetime.fromisoformat(ref.replace("Z", "+00:00")) if isinstance(ref, str) else ref
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days_open = (now - dt).days
            except Exception:
                days_open = None
        o["days_open"] = days_open
        o["is_overdue"] = bool(o.get("status") == "Pending" and days_open is not None and days_open >= threshold)
    return items


@api_router.patch("/orders/{oid}/status")
async def update_order_status(oid: str, body: OrderStatusUpdate, user=Depends(get_current_user)):
    if body.status not in ("Pending", "Dispatched", "Cleared"):
        raise HTTPException(status_code=400, detail="Invalid status")
    res = await db.orders.update_one({"id": oid}, {"$set": {"status": body.status, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api_router.patch("/orders/{oid}")
async def admin_update_order(oid: str, body: OrderUpdate, admin=Depends(require_admin)):
    """Admin-only full edit of an order: customer, items, dates, notes, status."""
    existing = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Order not found")
    update: Dict[str, Any] = {}
    if body.customer_id is not None and body.customer_id != existing.get("customer_id"):
        cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        update["customer_id"] = body.customer_id
        update["customer_name"] = cust["name"]
    if body.items is not None:
        if not body.items:
            raise HTTPException(status_code=400, detail="Order must contain at least one item")
        item_ids = [it.item_id for it in body.items]
        found = await db.items.find({"id": {"$in": item_ids}}, {"_id": 0, "id": 1}).to_list(1000)
        found_ids = {f["id"] for f in found}
        missing = [iid for iid in item_ids if iid not in found_ids]
        if missing:
            raise HTTPException(status_code=400, detail=f"Unknown item_id(s): {missing}")
        update["items"] = [it.model_dump() for it in body.items]
    if body.order_date is not None:
        update["order_date"] = body.order_date
    if body.delivery_date is not None:
        update["delivery_date"] = body.delivery_date
    if body.notes is not None:
        update["notes"] = body.notes
    if body.status is not None:
        if body.status not in ("Pending", "Dispatched", "Cleared"):
            raise HTTPException(status_code=400, detail="Invalid status")
        update["status"] = body.status
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = now_iso()
    await db.orders.update_one({"id": oid}, {"$set": update})
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api_router.delete("/orders/{oid}")
async def delete_order(oid: str, user=Depends(require_admin)):
    res = await db.orders.delete_one({"id": oid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"ok": True}


# ======================== Dashboard Summary ========================
@api_router.get("/dashboard/summary")
async def dashboard_summary(user=Depends(get_current_user)):
    pending = await db.orders.find({"status": "Pending"}, {"_id": 0}).to_list(5000)

    # Read overdue threshold (default 15) so the dashboard can rank overdue customers.
    settings_doc = await _get_settings_doc()
    overdue_days_threshold = int(settings_doc.get("overdue_days", 15))
    now = datetime.now(timezone.utc)

    # Strict item-wise aggregation. Key = item_id (fallback to item_name for
    # legacy rows that may exist before the strict requirement landed).
    item_totals: Dict[str, Dict[str, Any]] = {}
    party_breakdown: Dict[str, Dict[str, Dict[str, Any]]] = {}
    # customer_id -> aggregated overdue stats
    overdue_customers: Dict[str, Dict[str, Any]] = {}
    for o in pending:
        cust_name = o["customer_name"]
        cust_id = o.get("customer_id")
        order_id = o.get("id")
        order_date = o.get("order_date") or o.get("created_at")
        party_breakdown.setdefault(cust_name, {})

        # days_open from order_date
        days_open = 0
        odt = o.get("order_date") or o.get("created_at")
        if isinstance(odt, str):
            try:
                dt = datetime.fromisoformat(odt.replace("Z", "+00:00"))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days_open = max(0, (now - dt).days)
            except Exception:
                days_open = 0
        elif isinstance(odt, datetime):
            dt = odt if odt.tzinfo else odt.replace(tzinfo=timezone.utc)
            days_open = max(0, (now - dt).days)

        line_total = 0
        for it in o.get("items", []):
            # Robust qty parsing — strip stray commas/spaces from string-typed
            # legacy values so the grand total is always an integer sum.
            raw_qty = it.get("quantity") or 0
            if isinstance(raw_qty, str):
                raw_qty = raw_qty.replace(",", "").replace(" ", "").strip() or 0
            try:
                qty = int(float(raw_qty))
            except (TypeError, ValueError):
                qty = 0
            line_total += qty
            iid = it.get("item_id") or f"legacy:{it.get('item_name') or it.get('product_name')}"
            iname = it.get("item_name") or it.get("product_name") or "Unknown"
            pname = it.get("product_name") or ""

            row = item_totals.setdefault(iid, {
                "item_id": iid, "item_name": iname,
                "product_name": pname, "quantity": 0,
                "order_count": 0, "breakdown": [],
            })
            row["quantity"] += qty
            row["order_count"] += 1
            row["breakdown"].append({
                "order_id": order_id,
                "customer_name": cust_name,
                "quantity": qty,
                "order_date": order_date,
            })

            prow = party_breakdown[cust_name].setdefault(iid, {
                "item_id": iid, "item_name": iname,
                "product_name": pname, "quantity": 0,
            })
            prow["quantity"] += qty

        # Aggregate overdue stats per customer (only if past threshold).
        if days_open >= overdue_days_threshold:
            key = cust_id or f"name:{cust_name}"
            entry = overdue_customers.setdefault(key, {
                "customer_id": cust_id,
                "customer_name": cust_name,
                "oldest_days": 0,
                "pending_count": 0,
                "total_pcs": 0,
            })
            entry["oldest_days"] = max(entry["oldest_days"], days_open)
            entry["pending_count"] += 1
            entry["total_pcs"] += line_total

    total_orders = await db.orders.count_documents({})
    pending_count = len(pending)
    dispatched_count = await db.orders.count_documents({"status": "Dispatched"})
    cleared_count = await db.orders.count_documents({"status": "Cleared"})
    customers_count = await db.customers.count_documents({})
    products_count = await db.products.count_documents({})

    # Sort breakdown by quantity desc within each SKU row.
    for row in item_totals.values():
        row["breakdown"].sort(key=lambda x: -x["quantity"])

    item_totals_list = sorted(item_totals.values(), key=lambda x: -x["quantity"])
    party_list = [
        {"customer_name": c, "items": sorted(items.values(), key=lambda x: -x["quantity"])}
        for c, items in party_breakdown.items()
    ]
    overdue_customers_list = sorted(
        overdue_customers.values(),
        key=lambda x: (-x["oldest_days"], -x["pending_count"]),
    )

    return {
        "stats": {
            "total_orders": total_orders,
            "pending_orders": pending_count,
            "dispatched_orders": dispatched_count,
            "cleared_orders": cleared_count,
            "customers": customers_count,
            "products": products_count,
        },
        # Item-wise pending totals (strict requirement)
        "item_totals": item_totals_list,
        # Kept for backward compat — same item-wise rows
        "product_totals": item_totals_list,
        "party_breakdown": party_list,
        "overdue_customers": overdue_customers_list,
        "overdue_threshold_days": overdue_days_threshold,
    }


# ======================== Dispatch Matching ========================
@api_router.post("/dispatch/match")
async def dispatch_match(body: DispatchStockIn, user=Depends(get_current_user)):
    """Strict item-wise dispatch matching.

    Input: `items` dict where keys are **item_id** (SKU) and values are qty
    available. Suggests which pending orders can be fulfilled, allocating by
    item_id. Bag calculation is grouped per master product (since bags are
    packed per master product, not per SKU)."""
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    pmap = {p["name"]: p for p in products}

    # Look up item metadata for everything in the stock input so we can
    # render the response with proper item_name/product_name.
    stock_item_ids = [k for k, v in body.items.items() if int(v) > 0]
    item_docs = await db.items.find({"id": {"$in": stock_item_ids}}, {"_id": 0}).to_list(1000)
    imap = {it["id"]: it for it in item_docs}

    pending = await db.orders.find({"status": "Pending"}, {"_id": 0}).sort("created_at", 1).to_list(5000)

    # remaining stock map, keyed by item_id
    remaining: Dict[str, int] = {k: int(v) for k, v in body.items.items() if int(v) > 0}

    suggestions = []  # per order
    per_item_allocated: Dict[str, int] = {}
    per_product_allocated: Dict[str, int] = {}  # master product → qty (for bag calc)

    for o in pending:
        order_alloc = []
        any_fulfilled = False
        for it in o.get("items", []):
            iid = it.get("item_id")
            iname = it.get("item_name") or it.get("product_name") or "Unknown"
            pn = it.get("product_name") or ""
            need = int(it.get("quantity") or 0)
            avail = remaining.get(iid, 0) if iid else 0
            give = min(need, avail)
            if give > 0 and iid:
                remaining[iid] = avail - give
                per_item_allocated[iid] = per_item_allocated.get(iid, 0) + give
                per_product_allocated[pn] = per_product_allocated.get(pn, 0) + give
                any_fulfilled = True
            order_alloc.append({
                "item_id": iid,
                "item_name": iname,
                "product_name": pn,
                "needed": need,
                "allocated": give,
                "shortfall": need - give,
                "variant": it.get("variant"),
                "fully_fulfilled": give == need and need > 0,
            })
        if any_fulfilled:
            suggestions.append({
                "order_id": o["id"],
                "customer_id": o["customer_id"],
                "customer_name": o["customer_name"],
                "order_date": o.get("order_date"),
                "delivery_date": o.get("delivery_date"),
                "allocations": order_alloc,
                "fully_fulfilled": all(a["fully_fulfilled"] or a["needed"] == 0 for a in order_alloc),
            })

    # Bag calculation: per-SKU when the item has its own override
    # (item.min_per_bag / item.max_per_bag), else grouped per master product.
    # Bags never mix products. SKUs with custom bag size pack separately.
    bag_calc = []
    product_residual: Dict[str, int] = {}  # qty per master product (no item override)
    for iid, qty in per_item_allocated.items():
        meta = imap.get(iid, {})
        i_min = meta.get("min_per_bag")
        i_max = meta.get("max_per_bag")
        if i_min and i_max and i_min > 0 and i_max > 0:
            min_bags = -(-qty // i_max)
            max_bags = -(-qty // i_min)
            bag_calc.append({
                "scope": "item",
                "item_id": iid,
                "item_name": meta.get("name", "Unknown"),
                "product_name": meta.get("product_name", ""),
                "allocated_qty": qty,
                "min_per_bag": i_min,
                "max_per_bag": i_max,
                "min_bags": min_bags,
                "max_bags": max_bags,
                "bag_range_label": f"{min_bags} bag{'s' if min_bags != 1 else ''}" if min_bags == max_bags else f"{min_bags}–{max_bags} bags",
            })
        else:
            pn = meta.get("product_name", "")
            product_residual[pn] = product_residual.get(pn, 0) + qty

    for pn, qty in product_residual.items():
        prod = pmap.get(pn)
        if prod:
            min_b = prod.get("min_per_bag") or 1
            max_b = prod.get("max_per_bag") or min_b
            min_bags = -(-qty // max_b) if max_b > 0 else 0  # ceil
            max_bags = -(-qty // min_b) if min_b > 0 else 0
            bag_calc.append({
                "scope": "product",
                "product_name": pn,
                "allocated_qty": qty,
                "min_per_bag": min_b,
                "max_per_bag": max_b,
                "min_bags": min_bags,
                "max_bags": max_bags,
                "bag_range_label": f"{min_bags} bag{'s' if min_bags != 1 else ''}" if min_bags == max_bags else f"{min_bags}–{max_bags} bags",
            })
        else:
            bag_calc.append({"scope": "product", "product_name": pn, "allocated_qty": qty, "min_bags": 0, "max_bags": 0, "bag_range_label": "N/A"})

    leftover = []
    for iid, qty in remaining.items():
        if qty <= 0:
            continue
        meta = imap.get(iid, {})
        leftover.append({
            "item_id": iid,
            "item_name": meta.get("name", "Unknown"),
            "product_name": meta.get("product_name", ""),
            "quantity": qty,
        })

    return {
        "suggestions": suggestions,
        "bag_calculation": bag_calc,
        "leftover_stock": leftover,
        "per_item_allocated": [
            {
                "item_id": iid,
                "item_name": imap.get(iid, {}).get("name", "Unknown"),
                "product_name": imap.get(iid, {}).get("product_name", ""),
                "allocated_qty": qty,
            }
            for iid, qty in per_item_allocated.items()
        ],
        "input_stock": body.items,
    }


@api_router.post("/dispatch/execute")
async def dispatch_execute(body: DispatchExecuteIn, user=Depends(get_current_user)):
    """Partially fulfill a pending order. Subtracts the given quantities from
    each item line. Items hitting 0 are removed. If the order has no items
    left, it is marked Dispatched; otherwise it stays Pending so the
    remaining lot can be shipped later. A dispatch history record is saved."""
    order = await db.orders.find_one({"id": body.order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") != "Pending":
        raise HTTPException(status_code=400, detail=f"Order is {order.get('status')}, not Pending")
    if not body.allocations:
        raise HTTPException(status_code=400, detail="No allocations supplied")

    # Index order items by item_id for quick lookup
    order_items = order.get("items", [])
    by_iid: Dict[str, Dict[str, Any]] = {}
    for it in order_items:
        iid = it.get("item_id")
        if iid:
            by_iid[iid] = it

    # Validate allocations + price lookup against customer's assigned price list
    cust = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0}) or {}
    cust_price_list = cust.get("price_list_id")
    cust_transport = cust.get("transport_name") or ""

    dispatched_lines: List[Dict[str, Any]] = []
    for alloc in body.allocations:
        if alloc.quantity <= 0:
            continue
        line = by_iid.get(alloc.item_id)
        if not line:
            raise HTTPException(status_code=400, detail=f"Item {alloc.item_id} not in this order")
        remaining = int(line.get("quantity") or 0)
        if alloc.quantity > remaining:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot dispatch {alloc.quantity} of '{line.get('item_name')}' — only {remaining} remaining in order",
            )
        pricing = await compute_line_pricing(
            cust_price_list,
            alloc.item_id,
            line.get("product_name") or "",
        )
        dispatched_lines.append({
            "item_id": alloc.item_id,
            "item_name": line.get("item_name") or line.get("product_name"),
            "product_name": line.get("product_name"),
            "variant": line.get("variant"),
            "quantity": alloc.quantity,
            **pricing,
        })

    if not dispatched_lines:
        raise HTTPException(status_code=400, detail="All allocation quantities are zero")

    # Apply subtraction: keep only lines with qty > 0
    new_items: List[Dict[str, Any]] = []
    dispatched_map = {d["item_id"]: d["quantity"] for d in dispatched_lines}
    for it in order_items:
        iid = it.get("item_id")
        give = dispatched_map.get(iid, 0)
        new_qty = int(it.get("quantity") or 0) - give
        if new_qty > 0:
            new_items.append({**it, "quantity": new_qty})
        # else: line fully dispatched, drop it

    new_status = "Dispatched" if not new_items else "Pending"

    update_doc = {
        "items": new_items,
        "status": new_status,
        "updated_at": now_iso(),
    }
    # Preserve original item list on first partial dispatch for traceability
    if new_status == "Pending" and "original_items" not in order:
        update_doc["original_items"] = order_items

    await db.orders.update_one({"id": body.order_id}, {"$set": update_doc})

    # Save dispatch history record — MERGE into today's existing slip for
    # this customer if one exists, so multiple same-day dispatches to one
    # party produce ONE consolidated slip instead of many.
    # NOTE: total_value (the printed Bill Amount) is NEVER auto-computed
    # from item pricing — it must be entered manually by the operator in
    # Daily Report / Dispatch Ledger edit. New dispatches start at 0.
    existing = await _find_open_dispatch_today(order["customer_id"], order["customer_name"])
    if existing:
        merged_items = _merge_dispatch_lines(existing.get("items") or [], dispatched_lines)
        merged_total_pcs = sum(int(it.get("quantity") or 0) for it in merged_items)
        # Track all parent orders contributing to this slip (for traceability)
        order_ids = list(existing.get("order_ids") or [])
        if existing.get("order_id") and existing["order_id"] not in order_ids:
            order_ids.append(existing["order_id"])
        if body.order_id and body.order_id not in order_ids:
            order_ids.append(body.order_id)
        merged_notes = (existing.get("notes") or "").strip()
        if body.notes:
            merged_notes = (merged_notes + " | " + body.notes.strip()).strip(" |") if merged_notes else body.notes.strip()
        update_set: Dict[str, Any] = {
            "items": merged_items,
            "total_pcs": merged_total_pcs,
            # total_value is preserved — operator-entered bill amount is never overwritten
            "order_ids": order_ids,
            "notes": merged_notes,
            "last_dispatched_at": now_iso(),
            "last_dispatched_by": user["email"],
        }
        # Once any line fully clears its order, propagate the "fully" flag up
        if new_status == "Dispatched":
            update_set["order_fully_dispatched"] = True
        await db.dispatches.update_one({"id": existing["id"]}, {"$set": update_set})
        dispatch_doc = await db.dispatches.find_one({"id": existing["id"]}, {"_id": 0})
    else:
        dispatch_doc = {
            "id": str(uuid.uuid4()),
            "slip_no": await next_slip_no(),
            "order_id": body.order_id,
            "order_ids": [body.order_id],
            "customer_id": order["customer_id"],
            "customer_name": order["customer_name"],
            "transport_name": cust_transport,
            "price_list_id": cust_price_list,
            "items": dispatched_lines,
            "total_pcs": sum(d["quantity"] for d in dispatched_lines),
            "total_value": 0,  # operator must fill in Daily Report / edit
            "notes": body.notes or "",
            "dispatched_by": user["email"],
            "dispatched_at": now_iso(),
            "order_fully_dispatched": new_status == "Dispatched",
        }
        await db.dispatches.insert_one(dispatch_doc)
        dispatch_doc.pop("_id", None)

    updated_order = await db.orders.find_one({"id": body.order_id}, {"_id": 0})
    return {
        "dispatch": dispatch_doc,
        "order": updated_order,
        "fully_dispatched": new_status == "Dispatched",
    }


async def next_slip_no() -> int:
    """Atomically increment and return the next sequential dispatch slip number."""
    doc = await db.counters.find_one_and_update(
        {"_id": "dispatch_slip"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,  # ReturnDocument.AFTER
    )
    # `return_document=True` returns the post-update doc on Motor
    return int(doc["seq"]) if doc else 1


async def _find_open_dispatch_today(
    customer_id: Optional[str], customer_name: Optional[str]
) -> Optional[Dict[str, Any]]:
    """Find an existing dispatch slip created on the SAME business day (IST)
    for the same customer. Used to merge multiple same-day dispatches into a
    single slip. Matches by customer_id when present, else by customer_name
    (walk-in). India Standard Time (UTC+5:30) is used so the "day" boundary
    matches when the factory actually operates."""
    IST = timezone(timedelta(hours=5, minutes=30))
    today_ist = datetime.now(IST).date()
    # Window covering all of today (IST), expressed in UTC for the ISO-string
    # comparison against `dispatched_at` (which is stored in UTC).
    start = datetime.combine(today_ist, datetime.min.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    end = datetime.combine(today_ist, datetime.max.time(), tzinfo=IST).astimezone(timezone.utc).isoformat()
    q: Dict[str, Any] = {"dispatched_at": {"$gte": start, "$lte": end}}
    if customer_id:
        q["customer_id"] = customer_id
    else:
        q["customer_id"] = None
        q["customer_name"] = (customer_name or "").strip()
    return await db.dispatches.find_one(q, {"_id": 0}, sort=[("dispatched_at", 1)])


def _merge_dispatch_lines(
    existing: List[Dict[str, Any]], incoming: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Combine new dispatch lines into the existing list. Lines with the same
    item_id AND same net_unit_price are merged (quantities summed); otherwise
    the new line is appended as-is so different pricing stays auditable."""
    out: List[Dict[str, Any]] = [dict(x) for x in existing]
    for new in incoming:
        new_iid = new.get("item_id")
        new_net = float(new.get("net_unit_price") or 0)
        match_idx = None
        for i, ex in enumerate(out):
            if ex.get("item_id") == new_iid and float(ex.get("net_unit_price") or 0) == new_net:
                match_idx = i
                break
        if match_idx is not None:
            merged = dict(out[match_idx])
            merged["quantity"] = int(merged.get("quantity") or 0) + int(new.get("quantity") or 0)
            net = float(merged.get("net_unit_price") or 0)
            merged["line_value"] = round(net * merged["quantity"], 2)
            out[match_idx] = merged
        else:
            line = dict(new)
            if "line_value" not in line:
                line["line_value"] = round(
                    float(line.get("net_unit_price") or 0) * int(line.get("quantity") or 0), 2
                )
            out.append(line)
    return out


async def next_receipt_no() -> int:
    """Atomically increment and return the next sequential payment receipt number."""
    doc = await db.counters.find_one_and_update(
        {"_id": "payment_receipt"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return int(doc["seq"]) if doc else 1


@api_router.get("/dispatches")
async def list_dispatches(order_id: Optional[str] = None, customer_id: Optional[str] = None,
                          user=Depends(get_current_user)):
    q = {}
    if order_id:
        q["order_id"] = order_id
    if customer_id:
        q["customer_id"] = customer_id
    items = await db.dispatches.find(q, {"_id": 0}).sort("dispatched_at", -1).to_list(2000)
    return items


# ======================== Admin Dispatch Ledger (per-dispatch GR + slip) ========================
@api_router.get("/admin/dispatch-ledger")
async def admin_dispatch_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    limit: int = 200,
    skip: int = 0,
    user=Depends(get_current_user),
):
    """Dispatch ledger — every dispatch (regular + off-order) with its GR
    number. Accessible to any authenticated user so it can live in the main
    interface (not buried under Admin Settings). Newest first."""
    q: Dict[str, Any] = {}
    if start_date or end_date:
        rng: Dict[str, str] = {}
        if start_date:
            try:
                d = datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
            rng["$gte"] = datetime.combine(d, datetime.min.time(), tzinfo=timezone.utc).isoformat()
        if end_date:
            try:
                d = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="end_date must be YYYY-MM-DD")
            rng["$lte"] = datetime.combine(d, datetime.max.time(), tzinfo=timezone.utc).isoformat()
        q["dispatched_at"] = rng
    if customer_id:
        q["customer_id"] = customer_id
    cursor = (
        db.dispatches.find(q, {"_id": 0})
        .sort("dispatched_at", -1)
        .skip(max(0, int(skip)))
        .limit(max(1, min(500, int(limit))))
    )
    rows = await cursor.to_list(length=500)
    total = await db.dispatches.count_documents(q)
    grand_value = round(sum(float(r.get("total_value") or 0) for r in rows), 2)
    grand_pcs = sum(int(r.get("total_pcs") or 0) for r in rows)
    return {"total": total, "items": rows, "grand_total_value": grand_value, "grand_total_pcs": grand_pcs}


@api_router.patch("/dispatches/{did}/gr")
async def update_dispatch_gr(did: str, body: DispatchGrUpdate, user=Depends(get_current_user)):
    """Set / update the GR (Goods Receipt) number for a dispatch.
    Any authenticated user can edit this field (per product spec)."""
    gr = (body.gr_number or "").strip()
    res = await db.dispatches.update_one(
        {"id": did},
        {"$set": {
            "gr_number": gr,
            "gr_updated_at": now_iso(),
            "gr_updated_by": user["email"],
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    return await db.dispatches.find_one({"id": did}, {"_id": 0})


# ---- Dispatch row edit / delete (used by Single Party Ledger row actions) ----


class DispatchEditItem(BaseModel):
    item_id: Optional[str] = None
    item_name: str
    product_name: Optional[str] = ""
    variant: Optional[str] = ""
    quantity: float
    unit_price: Optional[float] = 0.0
    net_unit_price: Optional[float] = None
    discount_value: Optional[float] = 0.0
    discount_type: Optional[str] = ""


class DispatchEdit(BaseModel):
    gr_number: Optional[str] = None
    transport_name: Optional[str] = None
    notes: Optional[str] = None
    total_value: Optional[float] = None  # allow overriding bill amount (debit)
    bag_count: Optional[int] = None  # operator-entered number of bags shipped
    items: Optional[List[DispatchEditItem]] = None  # if provided, replaces line items


@api_router.patch("/dispatches/{did}")
async def update_dispatch(did: str, body: DispatchEdit, user=Depends(get_current_user)):
    """Edit a dispatch's bookkeeping fields (GR, transport, notes, bill
    amount, bag count) and optionally its line items (name / qty / price).
    Editing items here is a pure ledger correction — it does NOT rebalance
    the parent order's pending quantities. Any authenticated user."""
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": user["email"]}
    if body.gr_number is not None:
        upd["gr_number"] = body.gr_number.strip()
    if body.transport_name is not None:
        upd["transport_name"] = body.transport_name.strip()
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if body.bag_count is not None:
        if int(body.bag_count) < 0:
            raise HTTPException(status_code=400, detail="bag_count cannot be negative")
        upd["bag_count"] = int(body.bag_count)
    if body.items is not None:
        new_items: List[Dict[str, Any]] = []
        for it in body.items:
            qty = int(round(float(it.quantity or 0)))
            if qty <= 0 or not (it.item_name or "").strip():
                # skip empty / invalid rows silently
                continue
            unit = float(it.unit_price or 0)
            net = float(it.net_unit_price) if it.net_unit_price is not None else unit
            new_items.append({
                "item_id": it.item_id or str(uuid.uuid4()),
                "item_name": it.item_name.strip(),
                "product_name": (it.product_name or "").strip(),
                "variant": (it.variant or "").strip(),
                "quantity": qty,
                "unit_price": round(unit, 2),
                "discount_value": round(float(it.discount_value or 0), 2),
                "discount_type": (it.discount_type or ""),
                "net_unit_price": round(net, 2),
                "line_value": round(net * qty, 2),
            })
        if not new_items:
            raise HTTPException(status_code=400, detail="At least one item with name and quantity > 0 is required")
        upd["items"] = new_items
        upd["total_pcs"] = sum(it["quantity"] for it in new_items)
        # NOTE: total_value (Bill Amount) is NEVER auto-recomputed from
        # items — the operator must set it explicitly via the override field.
        # Existing value is preserved on item-only edits.
    if body.total_value is not None:
        if float(body.total_value) < 0:
            raise HTTPException(status_code=400, detail="total_value cannot be negative")
        upd["total_value"] = round(float(body.total_value), 2)
    res = await db.dispatches.update_one({"id": did}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    return await db.dispatches.find_one({"id": did}, {"_id": 0})


@api_router.delete("/dispatches/{did}")
async def delete_dispatch(did: str, admin=Depends(require_admin)):
    """Delete a dispatch record (admin only).

    NOTE: This removes the dispatch entry from the party ledger. It does NOT
    automatically restore the parent order's pending quantities — restoring
    inventory is a separate, more careful operation. Treat this as a
    bookkeeping correction for wrongly-entered ledger lines.
    """
    res = await db.dispatches.delete_one({"id": did})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    return {"ok": True, "deleted": did}


@api_router.get("/admin/dispatch-ledger/{did}/slip")
async def admin_dispatch_slip(did: str, user=Depends(get_current_user)):
    """Full slip payload (customer details + lines + totals) used by the
    frontend to render a printable dispatch slip. Accessible to any
    authenticated user."""
    d = await db.dispatches.find_one({"id": did}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    cust: Dict[str, Any] = {}
    if d.get("customer_id"):
        cust = await db.customers.find_one({"id": d["customer_id"]}, {"_id": 0}) or {}
    return {
        "dispatch": d,
        "customer": {
            "id": cust.get("id"),
            "name": d.get("customer_name") or cust.get("name") or "—",
            "phone": cust.get("phone") or "",
            "address": cust.get("address") or "",
            "city": cust.get("city") or "",
            "location": cust.get("location") or "",
            "transport_name": d.get("transport_name") or cust.get("transport_name") or "",
            "private_mark": cust.get("private_mark") or "",
        },
    }



# ======================== Price Lists ========================
async def compute_line_pricing(price_list_id: Optional[str], item_id: str, product_name: str) -> Dict[str, Any]:
    """Resolve unit_price, discount_value, discount_type and net_unit_price
    for one line, given an optional price-list assignment.
    If no price list / no entry exists, returns 0 values gracefully."""
    out = {
        "unit_price": 0.0,
        "discount_value": 0.0,
        "discount_type": "",
        "net_unit_price": 0.0,
    }
    if not price_list_id:
        return out
    pli = await db.price_list_items.find_one(
        {"price_list_id": price_list_id, "item_id": item_id}, {"_id": 0}
    )
    if pli:
        out["unit_price"] = float(pli.get("price") or 0)
    disc = await db.price_list_category_discounts.find_one(
        {"price_list_id": price_list_id, "product_name": product_name}, {"_id": 0}
    )
    if disc:
        out["discount_value"] = float(disc.get("discount_value") or 0)
        out["discount_type"] = disc.get("discount_type") or ""
    # Defensive: if a discount value is set but type is missing (legacy /
    # half-saved data), treat it as flat rupees off so the discount
    # actually applies instead of being silently ignored.
    if out["discount_value"] > 0 and out["discount_type"] not in ("₹", "%"):
        out["discount_type"] = "₹"
    # Net unit price after discount
    net = out["unit_price"]
    if out["discount_type"] == "%":
        net = max(0.0, net * (1 - out["discount_value"] / 100.0))
    elif out["discount_type"] == "₹":
        net = max(0.0, net - out["discount_value"])
    out["net_unit_price"] = round(net, 2)
    return out


@api_router.get("/price-lists")
async def list_price_lists(user=Depends(get_current_user)):
    lists = await db.price_lists.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    # attach counts
    out = []
    for pl in lists:
        cnt = await db.price_list_items.count_documents({"price_list_id": pl["id"]})
        dcnt = await db.price_list_category_discounts.count_documents({"price_list_id": pl["id"]})
        out.append({**pl, "items_count": cnt, "discounts_count": dcnt})
    return out


@api_router.post("/price-lists")
async def create_price_list(body: PriceListIn, admin=Depends(require_admin)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    if await db.price_lists.find_one({"name": name}):
        raise HTTPException(status_code=400, detail="A price list with this name already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "description": (body.description or "").strip(),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.price_lists.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/price-lists/{plid}")
async def update_price_list(plid: str, body: PriceListUpdate, admin=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    if "name" in update:
        update["name"] = update["name"].strip()
    update["updated_at"] = now_iso()
    res = await db.price_lists.update_one({"id": plid}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Price list not found")
    return await db.price_lists.find_one({"id": plid}, {"_id": 0})


@api_router.delete("/price-lists/{plid}")
async def delete_price_list(plid: str, admin=Depends(require_admin)):
    existing = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Price list not found")
    refs = await db.customers.count_documents({"price_list_id": plid})
    if refs > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete: {refs} customer(s) are assigned to this price list")
    await db.price_lists.delete_one({"id": plid})
    await db.price_list_items.delete_many({"price_list_id": plid})
    await db.price_list_category_discounts.delete_many({"price_list_id": plid})
    return {"ok": True, "deleted": plid}


@api_router.get("/price-lists/{plid}")
async def get_price_list_detail(plid: str, user=Depends(get_current_user)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    items = await db.items.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    pli_list = await db.price_list_items.find({"price_list_id": plid}, {"_id": 0}).to_list(10000)
    price_map = {p["item_id"]: float(p.get("price") or 0) for p in pli_list}
    rows = []
    for it in items:
        rows.append({
            "item_id": it["id"],
            "item_name": it["name"],
            "product_name": it.get("product_name") or "",
            "price": price_map.get(it["id"], 0.0),
        })
    discounts = await db.price_list_category_discounts.find({"price_list_id": plid}, {"_id": 0}).to_list(500)
    return {"price_list": pl, "items": rows, "discounts": discounts}


@api_router.post("/price-lists/{plid}/items")
async def set_price_list_item(plid: str, body: PriceListItemIn, admin=Depends(require_admin)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    item = await db.items.find_one({"id": body.item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    doc = {
        "price_list_id": plid,
        "item_id": body.item_id,
        "item_name": item["name"],
        "product_name": item.get("product_name") or "",
        "price": float(body.price or 0),
        "updated_at": now_iso(),
    }
    await db.price_list_items.update_one(
        {"price_list_id": plid, "item_id": body.item_id},
        {"$set": doc},
        upsert=True,
    )
    return doc


@api_router.post("/price-lists/{plid}/discounts")
async def set_price_list_category_discount(plid: str, body: CategoryDiscountIn, admin=Depends(require_admin)):
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    if body.discount_type not in ("₹", "%", ""):
        raise HTTPException(status_code=400, detail="discount_type must be '₹', '%' or ''")
    doc = {
        "price_list_id": plid,
        "product_name": body.product_name,
        "discount_value": float(body.discount_value or 0),
        "discount_type": body.discount_type,
        "updated_at": now_iso(),
    }
    await db.price_list_category_discounts.update_one(
        {"price_list_id": plid, "product_name": body.product_name},
        {"$set": doc},
        upsert=True,
    )
    return doc


@api_router.get("/price-lists/{plid}/export")
async def export_price_list(plid: str, user=Depends(get_current_user)):
    """Download Excel with two columns: Item Name | Price (₹)."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    items = await db.items.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    pli_list = await db.price_list_items.find({"price_list_id": plid}, {"_id": 0}).to_list(10000)
    price_map = {p["item_id"]: float(p.get("price") or 0) for p in pli_list}
    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"
    ws.append(["Item Name", "Price (Rs)"])
    for it in items:
        ws.append([it["name"], price_map.get(it["id"], 0.0)])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() else "_" for c in pl["name"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="price_list_{safe_name}.xlsx"'},
    )


@api_router.post("/price-lists/{plid}/import")
async def import_price_list(plid: str, file: UploadFile = File(...), admin=Depends(require_admin)):
    """Upload Excel with rows: Item Name | Price. Matches each row to an
    existing item (case-insensitive, fuzzy fallback ≥85). Unknown rows are
    returned for review. Existing prices are overwritten."""
    from openpyxl import load_workbook
    pl = await db.price_lists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(status_code=404, detail="Price list not found")
    blob = await file.read()
    if not blob:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")
    ws = wb.active
    items = await db.items.find({}, {"_id": 0}).to_list(5000)
    by_lower = {it["name"].strip().lower(): it for it in items}
    item_names = [it["name"] for it in items]
    updated = 0
    unknown: List[Dict[str, Any]] = []
    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    # Detect header — if first cell is non-numeric and looks like "Item Name"-ish, skip it
    def is_header(r):
        if not r or len(r) < 2:
            return False
        b = r[1]
        if isinstance(b, (int, float)):
            return False
        return True
    if first_row and not is_header(first_row):
        # Treat first row as data
        rows_to_process = [first_row]
    else:
        rows_to_process = []
    rows_to_process.extend(rows_iter)
    for r in rows_to_process:
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        price_raw = r[1] if len(r) > 1 else None
        try:
            price = float(price_raw) if price_raw is not None and str(price_raw).strip() != "" else 0.0
        except Exception:
            unknown.append({"item_name": name, "reason": f"invalid price '{price_raw}'"})
            continue
        # Match item by exact lowercase first
        match = by_lower.get(name.lower())
        if not match:
            # fuzzy fallback
            best = rf_process.extractOne(name, item_names, scorer=fuzz.WRatio, processor=rf_utils.default_process)
            if best and best[1] >= 85:
                match = next((it for it in items if it["name"] == best[0]), None)
        if not match:
            unknown.append({"item_name": name, "reason": "no matching SKU"})
            continue
        await db.price_list_items.update_one(
            {"price_list_id": plid, "item_id": match["id"]},
            {"$set": {
                "price_list_id": plid,
                "item_id": match["id"],
                "item_name": match["name"],
                "product_name": match.get("product_name") or "",
                "price": price,
                "updated_at": now_iso(),
            }},
            upsert=True,
        )
        updated += 1
    await db.price_lists.update_one({"id": plid}, {"$set": {"updated_at": now_iso()}})
    return {"updated": updated, "unknown_count": len(unknown), "unknown": unknown[:50]}


# ======================== Daily Dispatch Report ========================
@api_router.get("/reports/daily-dispatch")
async def daily_dispatch_report(date: Optional[str] = None, user=Depends(get_current_user)):
    """Consolidated end-of-day report grouped by party (customer).
    `date` is YYYY-MM-DD; defaults to today (server UTC date)."""
    today = datetime.now(timezone.utc).date()
    if date:
        try:
            target = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    else:
        target = today
    start = datetime.combine(target, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    end = datetime.combine(target, datetime.max.time(), tzinfo=timezone.utc).isoformat()
    dispatches = await db.dispatches.find(
        {"dispatched_at": {"$gte": start, "$lte": end}},
        {"_id": 0},
    ).sort("dispatched_at", 1).to_list(5000)
    # Group by customer
    groups: Dict[str, Dict[str, Any]] = {}
    customers_cache: Dict[str, Dict[str, Any]] = {}
    grand_pcs = 0
    grand_value = 0.0
    for d in dispatches:
        cid = d.get("customer_id") or "unknown"
        cust = customers_cache.get(cid)
        if cust is None and cid != "unknown":
            cust = await db.customers.find_one({"id": cid}, {"_id": 0}) or {}
            customers_cache[cid] = cust
        if cust is None:
            cust = {}
        g = groups.get(cid)
        if not g:
            g = {
                "customer_id": cid,
                "customer_name": d.get("customer_name") or cust.get("name") or "—",
                "transport_name": cust.get("transport_name") or "",
                "phone": cust.get("phone") or "",
                "address": cust.get("address") or "",
                "city": cust.get("city") or "",
                "location": cust.get("location") or "",
                "private_mark": cust.get("private_mark") or "",
                "lines": [],
                "dispatches": [],
                "total_pcs": 0,
                "total_value": 0.0,
                "dispatch_count": 0,
            }
            groups[cid] = g
        g["dispatch_count"] += 1
        g["dispatches"].append({
            "id": d.get("id"),
            "slip_no": d.get("slip_no"),
            "gr_number": d.get("gr_number") or "",
            "total_value": float(d.get("total_value") or 0),
            "total_pcs": int(d.get("total_pcs") or 0),
            "dispatched_at": d.get("dispatched_at"),
        })
        for line in d.get("items", []):
            qty = int(line.get("quantity") or 0)
            # Recompute pricing LIVE against the customer's currently-assigned
            # price list + discounts. This makes the report reflect any
            # discount changes / price list updates the user makes AFTER the
            # dispatch happened (instead of being frozen at dispatch time).
            cust_pl = (cust or {}).get("price_list_id")
            live = await compute_line_pricing(
                cust_pl,
                line.get("item_id"),
                line.get("product_name") or "",
            )
            unit_price = float(live["unit_price"] or line.get("unit_price") or 0)
            discount_value = float(live["discount_value"] or 0)
            discount_type = live["discount_type"] or ""
            net_unit_price = float(live["net_unit_price"] or unit_price)
            # If the customer has no price list assigned now but the dispatch
            # snapshot has values, fall back to the snapshot so historical
            # rows aren't blanked out.
            if not cust_pl:
                unit_price = float(line.get("unit_price") or 0)
                discount_value = float(line.get("discount_value") or 0)
                discount_type = line.get("discount_type") or ""
                net_unit_price = float(line.get("net_unit_price") or unit_price)
            value = round(net_unit_price * qty, 2)
            g["lines"].append({
                "item_id": line.get("item_id"),
                "item_name": line.get("item_name"),
                "product_name": line.get("product_name"),
                "variant": line.get("variant"),
                "quantity": qty,
                "unit_price": unit_price,
                "discount_value": discount_value,
                "discount_type": discount_type,
                "net_unit_price": net_unit_price,
                "line_value": value,
                "dispatched_at": d.get("dispatched_at"),
                "dispatch_id": d.get("id"),
            })
            g["total_pcs"] += qty
            g["total_value"] += value
            grand_pcs += qty
            grand_value += value
        g["total_value"] = round(g["total_value"], 2)
    out_groups = sorted(groups.values(), key=lambda g: g["customer_name"].lower())
    return {
        "date": target.isoformat(),
        "groups": out_groups,
        "grand_total_pcs": grand_pcs,
        "grand_total_value": round(grand_value, 2),
        "dispatch_count": len(dispatches),
    }


# ======================== Voice Transcription ========================
@api_router.post("/voice/transcribe")
async def voice_transcribe(file: UploadFile = File(...), user=Depends(get_current_user)):
    # Validate input BEFORE attempting external service call — empty audio must
    # return 400, not 500, even when the LLM key is missing.
    audio_bytes = await file.read()
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="Empty audio file")
    try:
        from emergentintegrations.llm.openai import OpenAISpeechToText
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Speech library not available: {e}")
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="EMERGENT_LLM_KEY not configured")

    stt = OpenAISpeechToText(api_key=api_key)
    file_like = io.BytesIO(audio_bytes)
    file_like.name = file.filename or "audio.webm"
    # Build a biasing prompt that includes a sample of real party names so
    # Whisper transcribes them spelled correctly instead of phonetic noise.
    try:
        party_sample = await db.customers.find({}, {"_id": 0, "name": 1}).limit(60).to_list(60)
        party_names = ", ".join([c["name"] for c in party_sample if c.get("name")])
    except Exception:
        party_names = ""
    bias_prompt = (
        "Factory order for two-wheeler spare parts dictated in Hinglish / English / Hindi. "
        "Products: side stand, center stand with kit, center stand without kit, center stand pin, "
        "footrest rod, seat kunda, side seat handle, lady footrest, side footrest, handlebar, "
        "number plate, engine plate, v-bracket, luggage rod, side mirror clump, rear seat handle. "
        "Quantities in pieces, e.g. do sau, teen sau, char sau, paanch sau, ek hazaar."
    )
    if party_names:
        bias_prompt += f" Customer parties may include: {party_names}."
    try:
        resp = await stt.transcribe(
            file=file_like,
            model="whisper-1",
            response_format="json",
            prompt=bias_prompt,
            temperature=0.0,
        )
        text = getattr(resp, "text", None) or (resp.get("text") if isinstance(resp, dict) else None) or str(resp)
    except Exception as e:
        logger.exception("Transcription failed")
        raise HTTPException(status_code=500, detail=f"Transcription error: {e}")

    # Parse text into order items
    parsed = await parse_voice_order_with_items(text)
    customer_match = await match_customer_from_voice(text)
    return {"text": text, "parsed_items": parsed, "parsed_customer": customer_match}


async def match_customer_from_voice(text: str) -> Optional[Dict[str, Any]]:
    """Fuzzy-match a customer name out of a free-form voice transcript.

    The transcript usually starts with something like "<party name> ke liye ..."
    or "<party name> ko ..." or just dictates a list of items. We compare the
    *whole* transcript against every saved customer name using rapidfuzz's
    `token_set_ratio` (handles word reordering, partial mentions, transliteration
    noise from Whisper). A score ≥ 65 is treated as a confident hit so floor
    operators can correct the rest by ear.
    """
    if not text or not text.strip():
        return None
    customers = await db.customers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
    if not customers:
        return None
    names = [c["name"] for c in customers]
    lower_text = text.lower()
    lower_names = [n.lower() for n in names]
    # Strip product/SKU vocabulary so it doesn't drown the customer signal.
    stripped = lower_text
    for key in SYNONYM_MAP.keys():
        stripped = stripped.replace(key, " ")
    # token_set_ratio is robust to extra/unordered words. For longer party
    # names we also consider partial_ratio (handles "ramesh auto sons" being
    # mentioned as just "ramesh sons"), but only when the saved name is long
    # enough that partial_ratio won't trivially match random short substrings.
    best = None
    for i, n in enumerate(lower_names):
        s1 = fuzz.token_set_ratio(stripped, n)
        s2 = fuzz.partial_ratio(stripped, n) if len(n) >= 6 else 0
        score = max(s1, s2)
        if best is None or score > best[1]:
            best = (i, score)
    # 70 is the sweet spot: low enough to catch heavily-mangled Whisper output,
    # high enough that "xyzxyz random" doesn't false-positive on a 3-letter
    # party like "CAD".
    if not best or best[1] < 70:
        return None
    idx, score = best
    return {"id": customers[idx]["id"], "name": names[idx], "score": int(score)}


# Hinglish + English word-number lookups for voice-order quantity parsing.
_WORD_NUMS = {
    # English ones
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
    "eleven": 11, "twelve": 12, "fifteen": 15, "twenty": 20, "thirty": 30,
    "forty": 40, "fifty": 50, "sixty": 60, "seventy": 70, "eighty": 80, "ninety": 90,
    "hundred": 100, "thousand": 1000,
    # Hinglish ones / tens / hundreds
    "ek": 1, "do": 2, "teen": 3, "char": 4, "chaar": 4, "paanch": 5, "panch": 5,
    "che": 6, "chhe": 6, "saat": 7, "aath": 7, "nau": 9, "das": 10, "dus": 10,
    "bees": 20, "tees": 30, "chalis": 40, "pachas": 50,
    "saath": 60, "sattar": 70, "assi": 80, "nabbe": 90,
    "sau": 100, "hazar": 1000, "hazaar": 1000,
    # Common Whisper misspellings of Hinglish numbers
    "dosa": 200,        # "do sau" merged
    "doso": 200,
    "tinso": 300, "tinsau": 300, "teensau": 300,  # "teen sau"
    "charsau": 400, "chaarsau": 400, "charso": 400,
    "panchsau": 500, "paanchsau": 500, "paanchso": 500, "panchso": 500,
    "chesau": 600, "chhesau": 600, "cheso": 600, "chheso": 600,
    "saatsau": 700, "saatso": 700,
    "aathsau": 800, "aathso": 800,
    "nausau": 900, "nauso": 900,
    # Devanagari numerals (Whisper sometimes returns these for pure-Hindi input)
    "०": 0, "१": 1, "२": 2, "३": 3, "४": 4, "५": 5,
    "६": 6, "७": 7, "८": 8, "९": 9,
    # Hindi number words (Devanagari)
    "एक": 1, "दो": 2, "तीन": 3, "चार": 4, "पाँच": 5, "पांच": 5,
    "छह": 6, "छः": 6, "सात": 7, "आठ": 8, "नौ": 9, "दस": 10,
    "बीस": 20, "तीस": 30, "चालीस": 40, "पचास": 50,
    "साठ": 60, "सत्तर": 70, "अस्सी": 80, "नब्बे": 90,
    "सौ": 100, "हज़ार": 1000, "हजार": 1000,
}


def _parse_word_number(phrase: str) -> int:
    """Parse a small word-number phrase like 'do sau' (200), 'two hundred' (200),
    'teen sau pachas' (350), or pure Hindi 'दो सौ' (200). Falls back to 0 if
    nothing matches."""
    import re
    # Keep Latin letters AND Devanagari (\u0900-\u097F) so Hindi tokens survive
    # the cleanup. Hyphens become spaces so "two-hundred" → ["two", "hundred"].
    cleaned = re.sub(r"[^a-z\u0900-\u097F\s]", "", phrase.lower().replace("-", " "))
    tokens = [t for t in cleaned.split() if t]
    total = 0
    current = 0
    matched = False
    for t in tokens:
        if t not in _WORD_NUMS:
            continue
        matched = True
        v = _WORD_NUMS[t]
        if v >= 100:
            current = (current or 1) * v
            total += current
            current = 0
        else:
            current += v
    total += current
    return total if matched else 0


def parse_voice_order(text: str) -> List[Dict[str, Any]]:
    """Parse Hinglish voice transcript into [{product_name, quantity, matched_phrase}].

    Splits text by clause boundaries (, . ; ! ?) so a number can't bleed across
    clauses. Within each clause:
      - find non-overlapping product hits (longest synonym wins)
      - collect digit + Hinglish/English word-number candidates
      - assign each candidate to its NEAREST product hit; product takes its
        closest assigned candidate.
    """
    import re
    out: List[Dict[str, Any]] = []
    if not text:
        return out

    sorted_keys = sorted(SYNONYM_MAP.keys(), key=lambda x: -len(x))

    # Normalise whitespace and strip stray punctuation that breaks word
    # boundaries (Whisper sometimes injects double-quotes / Hindi danda etc).
    text = re.sub(r"[\"'`]+", " ", text)
    text = text.replace("।", ".")  # Hindi danda → period

    # Split into clauses
    clauses = re.split(r"[,.;!?\n]+", text.lower())

    for clause in clauses:
        if not clause.strip():
            continue
        # 1. find product hits in clause
        hits = []
        used_spans = []
        for key in sorted_keys:
            for m in re.finditer(re.escape(key), clause):
                s, e = m.span()
                if any(us <= s < ue or us < e <= ue for (us, ue) in used_spans):
                    continue
                used_spans.append((s, e))
                hits.append((s, e, key))
        if not hits:
            continue
        hits.sort(key=lambda x: x[0])

        # 2. collect quantity candidates within the clause
        def _in_product(pos: int) -> bool:
            return any(s <= pos < e for (s, e, _k) in hits)

        candidates: List[tuple] = []
        for m in re.finditer(r"\d{1,5}", clause):
            if not _in_product(m.start()):
                candidates.append((m.start(), int(m.group(0))))
        token_iter = list(re.finditer(r"[a-z\u0900-\u097F]+", clause))
        i_tok = 0
        while i_tok < len(token_iter):
            tok = token_iter[i_tok].group(0)
            if tok in _WORD_NUMS and not _in_product(token_iter[i_tok].start()):
                run_start = token_iter[i_tok].start()
                run_end = token_iter[i_tok].end()
                j = i_tok + 1
                while (
                    j < len(token_iter)
                    and token_iter[j].group(0) in _WORD_NUMS
                    and not _in_product(token_iter[j].start())
                ):
                    run_end = token_iter[j].end()
                    j += 1
                qty = _parse_word_number(clause[run_start:run_end])
                if qty > 0:
                    candidates.append((run_start, qty))
                i_tok = j
            else:
                i_tok += 1

        # 3. assign each candidate to nearest product hit using greedy 1-to-1
        # matching. Plain "nearest product" was incorrect: a long sentence like
        # "I need 500 side stands and three hundred footrest rods" causes BOTH
        # 500 and 300 to be distance-closer to side_stand than to footrest_rod,
        # leaving footrest_rod with qty 0. Greedy 1-to-1 fixes this: pair the
        # globally-shortest (candidate, product) edge first, then remove both
        # from the pool, then the next shortest, etc.
        def _dist(pos: int, s: int, e: int) -> int:
            if pos < s:
                return s - pos
            if pos >= e:
                return pos - e + 1
            return 0

        edges = []  # (distance, cand_index, hit_index, qty)
        for ci, (pos, qty) in enumerate(candidates):
            for hi, (s, e, _k) in enumerate(hits):
                edges.append((_dist(pos, s, e), ci, hi, qty))
        edges.sort(key=lambda x: x[0])

        product_qty: Dict[int, int] = {}
        used_cands: set = set()
        used_hits: set = set()
        for _d, ci, hi, qty in edges:
            if ci in used_cands or hi in used_hits:
                continue
            product_qty[hi] = qty
            used_cands.add(ci)
            used_hits.add(hi)
            if len(used_hits) == len(hits) or len(used_cands) == len(candidates):
                break

        for i, (s, e, key) in enumerate(hits):
            out.append({"product_name": SYNONYM_MAP[key], "quantity": product_qty.get(i, 0), "matched_phrase": key})

    return out


async def parse_voice_order_with_items(text: str) -> List[Dict[str, Any]]:
    """Wrap parse_voice_order() and try to match a specific item SKU within
    the surrounding text for each detected product hit. If no item match is
    strong enough, returns the row with item_id/item_name as None — the UI
    will prompt the user to pick an item before submit.
    """
    import re
    base = parse_voice_order(text)
    if not base:
        return base
    # Pull all items from DB once
    all_items = await db.items.find({}, {"_id": 0}).to_list(5000)
    by_product: Dict[str, List[Dict[str, Any]]] = {}
    for it in all_items:
        by_product.setdefault(it["product_name"], []).append(it)

    lower_text = (text or "").lower()
    for row in base:
        pname = row.get("product_name")
        candidates = by_product.get(pname) or []
        if not candidates:
            row["item_id"] = None
            row["item_name"] = None
            continue
        names = [c["name"] for c in candidates]
        # Use full text as query; rapidfuzz token_set/WRatio handles short queries vs long candidates
        match = rf_process.extractOne(lower_text, [n.lower() for n in names], scorer=fuzz.token_set_ratio)
        if match and match[1] >= 70:
            idx = match[2]
            chosen = candidates[idx]
            row["item_id"] = chosen["id"]
            row["item_name"] = chosen["name"]
            row["item_match_score"] = match[1]
        else:
            row["item_id"] = None
            row["item_name"] = None
    return base


# ======================== Login Attestation (security audit) ========================
# Stores an admin-only audit record per login containing the user's location
# and a webcam photo. Capture is consent-based: the frontend shows a clear
# notice and standard browser permission prompts. When the user declines,
# we still log the event with `consent=False` so admins see WHO declined.

MAX_PHOTO_BYTES = 600 * 1024  # cap stored base64 size at ~600 KB to avoid bloating the DB


def _client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for") or ""
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else ""


_MOBILE_UA_RE = re.compile(
    r"Android|webOS|iPhone|iPad|iPod|BlackBerry|IEMobile|Opera Mini|Mobi|Tablet",
    re.IGNORECASE,
)


def _is_mobile_ua(ua: str) -> bool:
    """Best-effort mobile/tablet detection from the User-Agent header."""
    if not ua:
        return False
    return bool(_MOBILE_UA_RE.search(ua))


@api_router.post("/auth/attestation")
async def create_login_attestation(
    body: LoginAttestationIn,
    request: Request,
    user=Depends(get_current_user),
):
    """Record a consent-based login security capture (photo + location)."""
    photo_b64 = body.photo_b64 or ""
    # Strip "data:image/jpeg;base64," prefix if present
    if photo_b64.startswith("data:") and "," in photo_b64:
        photo_b64 = photo_b64.split(",", 1)[1]
    if len(photo_b64) > MAX_PHOTO_BYTES * 4 // 3:  # base64 is ~33% larger than raw
        # Reject oversized payloads rather than silently truncating
        raise HTTPException(status_code=413, detail="Photo too large (max ~600 KB)")

    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "username": user.get("username") or user["email"],
        "user_name": user.get("name") or user.get("username") or user["email"],
        "role": user.get("role") or "user",
        "ip": _client_ip(request),
        "user_agent": request.headers.get("user-agent") or "",
        "is_mobile": _is_mobile_ua(request.headers.get("user-agent") or ""),
        "consent": bool(body.consent),
        "latitude": body.latitude,
        "longitude": body.longitude,
        "accuracy_meters": body.accuracy_meters,
        "has_photo": bool(photo_b64),
        "photo_b64": photo_b64 or None,
        "photo_skipped": bool(body.photo_skipped),
        "location_skipped": bool(body.location_skipped),
        "error": (body.error or "")[:500],
        "created_at": now_iso(),
    }
    await db.login_attestations.insert_one(doc)
    return {"id": doc["id"], "stored": True, "has_photo": doc["has_photo"]}


@api_router.get("/admin/login-attestations")
async def list_login_attestations(
    limit: int = 50,
    skip: int = 0,
    user_id: Optional[str] = None,
    consent: Optional[bool] = None,
    admin=Depends(require_admin),
):
    """Admin-only audit log of all login captures (most recent first)."""
    q: Dict[str, Any] = {}
    if user_id:
        q["user_id"] = user_id
    if consent is not None:
        q["consent"] = consent
    cursor = (
        db.login_attestations.find(q, {"_id": 0, "photo_b64": 0})
        .sort("created_at", -1)
        .skip(max(0, int(skip)))
        .limit(max(1, min(200, int(limit))))
    )
    rows = await cursor.to_list(length=200)
    total = await db.login_attestations.count_documents(q)
    return {"total": total, "items": rows}


@api_router.get("/admin/login-attestations/{att_id}/photo")
async def get_login_attestation_photo(att_id: str, admin=Depends(require_admin)):
    """Admin-only: stream the JPEG photo for a single attestation record."""
    rec = await db.login_attestations.find_one({"id": att_id}, {"_id": 0, "photo_b64": 1})
    if not rec or not rec.get("photo_b64"):
        raise HTTPException(status_code=404, detail="Photo not available")
    try:
        raw = base64.b64decode(rec["photo_b64"])
    except Exception:
        raise HTTPException(status_code=500, detail="Stored photo is not valid base64")
    return Response(content=raw, media_type="image/jpeg")


# ======================== Off-Order (Direct) Dispatch ========================
@api_router.post("/dispatch/off-order")
async def dispatch_off_order(body: OffOrderDispatchIn, admin=Depends(require_admin)):
    """Dispatch SKUs to a party with NO existing pending order.

    The created record lives in the `dispatches` collection alongside
    order-linked dispatches, so it is included in the Daily Dispatch
    Report automatically. `order_id` is `None` for these records.
    """
    # ---- Resolve customer (existing OR walk-in) ----
    cust: Dict[str, Any] = {}
    customer_id = body.customer_id or None
    customer_name = (body.customer_name or "").strip()
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0}) or {}
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_name = cust.get("name") or customer_name
    if not customer_name:
        raise HTTPException(status_code=400, detail="customer_id or customer_name is required")

    # ---- Validate items ----
    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    seen_ids: set = set()
    dispatched_lines: List[Dict[str, Any]] = []
    cust_price_list = cust.get("price_list_id")
    for line in body.items:
        if line.quantity is None or int(line.quantity) <= 0:
            raise HTTPException(status_code=400, detail="Each item quantity must be > 0")
        if line.item_id in seen_ids:
            raise HTTPException(status_code=400, detail=f"Duplicate item in payload: {line.item_id}")
        seen_ids.add(line.item_id)
        sku = await db.items.find_one({"id": line.item_id}, {"_id": 0})
        if not sku:
            raise HTTPException(status_code=404, detail=f"Item {line.item_id} not found")
        pricing = await compute_line_pricing(
            cust_price_list, line.item_id, sku.get("product_name") or ""
        )
        dispatched_lines.append({
            "item_id": line.item_id,
            "item_name": sku.get("name"),
            "product_name": sku.get("product_name"),
            "variant": sku.get("variant"),
            "quantity": int(line.quantity),
            **pricing,
        })

    # NOTE: total_value (Bill Amount) is NEVER auto-computed from item
    # pricing — operator must fill it manually in Daily Report. Item pricing
    # remains on each line for printable-slip reference only.

    # MERGE into today's existing slip for this customer if one exists, so
    # multiple same-day dispatches to one party produce ONE slip.
    existing = await _find_open_dispatch_today(customer_id, customer_name)
    if existing:
        merged_items = _merge_dispatch_lines(existing.get("items") or [], dispatched_lines)
        merged_total_pcs = sum(int(it.get("quantity") or 0) for it in merged_items)
        merged_notes = (existing.get("notes") or "").strip()
        if body.notes:
            extra = body.notes.strip()
            merged_notes = (merged_notes + " | " + extra).strip(" |") if merged_notes else extra
        await db.dispatches.update_one(
            {"id": existing["id"]},
            {"$set": {
                "items": merged_items,
                "total_pcs": merged_total_pcs,
                # total_value is preserved — operator-entered bill amount never overwritten
                "notes": merged_notes,
                "last_dispatched_at": now_iso(),
                "last_dispatched_by": admin["email"],
            }},
        )
        dispatch_doc = await db.dispatches.find_one({"id": existing["id"]}, {"_id": 0})
        return {"dispatch": dispatch_doc, "merged": True}

    dispatch_doc = {
        "id": str(uuid.uuid4()),
        "slip_no": await next_slip_no(),
        "order_id": None,  # explicit: no parent order
        "order_ids": [],
        "off_order": True,  # convenience flag for reporting/filtering
        "customer_id": customer_id,
        "customer_name": customer_name,
        "transport_name": (body.transport_name or cust.get("transport_name") or "").strip(),
        "price_list_id": cust_price_list,
        "items": dispatched_lines,
        "total_pcs": sum(d["quantity"] for d in dispatched_lines),
        "total_value": 0,  # operator must fill in Daily Report / edit
        "notes": (body.notes or "").strip(),
        "dispatched_by": admin["email"],
        "dispatched_at": now_iso(),
        "order_fully_dispatched": False,  # no order to "fully dispatch"
    }
    await db.dispatches.insert_one(dispatch_doc)
    dispatch_doc.pop("_id", None)
    return {"dispatch": dispatch_doc}


# ======================== Party Payments ========================
# Payments are the credit side of a party's ledger — money received from a
# customer (cash, UPI, NEFT/RTGS, cheque, etc.). Combined with dispatches
# (the debit side) they drive the running account balance shown in the
# Single Party Ledger page.

PAYMENT_SOURCES = {"cash", "upi", "bank_transfer", "neft", "rtgs", "cheque", "card", "adjustment", "other"}


class PaymentIn(BaseModel):
    customer_id: str
    amount: float
    source: str = "cash"
    reference: Optional[str] = ""  # UTR, cheque #, transaction id, etc.
    paid_at: Optional[str] = None  # ISO date or datetime; defaults to "now"
    notes: Optional[str] = ""
    # "cash" = received from customer (default).
    # "supplier_on_behalf" = we paid a 3rd-party supplier on the customer's
    # behalf; that amount credits the customer's ledger AND debits the
    # supplier's ledger (via a linked supplier_payment record).
    payment_mode: Optional[str] = "cash"
    paid_to_supplier_id: Optional[str] = None  # required when payment_mode = supplier_on_behalf


class PaymentUpdate(BaseModel):
    amount: Optional[float] = None
    source: Optional[str] = None
    reference: Optional[str] = None
    paid_at: Optional[str] = None
    notes: Optional[str] = None


def _normalize_payment_source(src: str) -> str:
    s = (src or "").strip().lower().replace("-", "_").replace(" ", "_")
    return s if s in PAYMENT_SOURCES else "other"


def _normalize_payment_dt(value: Optional[str]) -> str:
    """Accept '2026-06-12', '2026-06-12T15:30:00', or ISO with TZ. Always
    persist as ISO 8601 with seconds + timezone."""
    if not value:
        return datetime.now(timezone.utc).isoformat()
    try:
        s = value.strip()
        if len(s) == 10:  # bare YYYY-MM-DD
            return datetime.fromisoformat(s + "T00:00:00").replace(tzinfo=timezone.utc).isoformat()
        # full datetime — let fromisoformat parse
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat()
    except Exception:
        return datetime.now(timezone.utc).isoformat()


@api_router.post("/payments")
async def create_payment(body: PaymentIn, user=Depends(get_current_user)):
    if not body.customer_id:
        raise HTTPException(status_code=400, detail="customer_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    cust = await db.customers.find_one({"id": body.customer_id}, {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")

    mode = (body.payment_mode or "cash").strip().lower()
    supplier = None
    if mode == "supplier_on_behalf":
        if not body.paid_to_supplier_id:
            raise HTTPException(status_code=400, detail="paid_to_supplier_id required when payment_mode is supplier_on_behalf")
        supplier = await db.suppliers.find_one({"id": body.paid_to_supplier_id}, {"_id": 0})
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
    else:
        mode = "cash"

    pid = str(uuid.uuid4())
    payment = {
        "id": pid,
        "receipt_no": await next_receipt_no(),
        "customer_id": body.customer_id,
        "customer_name": cust.get("name") or "",
        "amount": round(float(body.amount), 2),
        "source": _normalize_payment_source(body.source),
        "reference": (body.reference or "").strip(),
        "paid_at": _normalize_payment_dt(body.paid_at),
        "notes": (body.notes or "").strip(),
        "payment_mode": mode,
        "paid_to_supplier_id": (body.paid_to_supplier_id or None) if mode == "supplier_on_behalf" else None,
        "paid_to_supplier_name": (supplier.get("name") if supplier else None),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.payments.insert_one(payment)
    payment.pop("_id", None)

    # If paid on behalf of customer to a supplier, mirror as a supplier_payment.
    if mode == "supplier_on_behalf" and supplier:
        sup_payment = {
            "id": str(uuid.uuid4()),
            "supplier_id": supplier["id"],
            "supplier_name": supplier.get("name") or "",
            "amount": payment["amount"],
            "source": payment["source"],
            "reference": payment["reference"],
            "paid_at": payment["paid_at"],
            "notes": payment["notes"],
            "on_behalf_of_customer_id": body.customer_id,
            "on_behalf_of_customer_name": cust.get("name") or "",
            "customer_payment_id": pid,
            "created_by": payment["created_by"],
            "created_at": payment["created_at"],
        }
        await db.supplier_payments.insert_one(sup_payment)

    return payment


@api_router.get("/payments")
async def list_payments(
    customer_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
    skip: int = 0,
    user=Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if customer_id:
        q["customer_id"] = customer_id
    if start_date or end_date:
        rng: Dict[str, Any] = {}
        if start_date:
            rng["$gte"] = start_date
        if end_date:
            rng["$lte"] = end_date + "T23:59:59"
        q["paid_at"] = rng
    cursor = db.payments.find(q, {"_id": 0}).sort("paid_at", -1).skip(skip).limit(limit)
    items = await cursor.to_list(length=limit)
    total = await db.payments.count_documents(q)
    total_amount = round(sum(float(p.get("amount") or 0) for p in items), 2)
    return {
        "items": items,
        "total": total,
        "total_amount": total_amount,
    }


@api_router.patch("/payments/{pid}")
async def update_payment(pid: str, body: PaymentUpdate, user=Depends(get_current_user)):
    existing = await db.payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Payment not found")
    upd: Dict[str, Any] = {}
    if body.amount is not None:
        if float(body.amount) <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than zero")
        upd["amount"] = round(float(body.amount), 2)
    if body.source is not None:
        upd["source"] = _normalize_payment_source(body.source)
    if body.reference is not None:
        upd["reference"] = body.reference.strip()
    if body.paid_at is not None:
        upd["paid_at"] = _normalize_payment_dt(body.paid_at)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    if upd:
        await db.payments.update_one({"id": pid}, {"$set": upd})
    return await db.payments.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/payments/{pid}")
async def delete_payment(pid: str, admin=Depends(require_admin)):
    # Cascade-remove the mirrored supplier_payment (if this was an on-behalf
    # payment) so the supplier ledger stays in sync.
    await db.supplier_payments.delete_many({"customer_payment_id": pid})
    res = await db.payments.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"ok": True}


# ======================== Suppliers (Admin) ========================
# Suppliers are vendors that provide raw materials. The supplier ledger
# mirrors the customer ledger: purchases (we owe them) are debits, payments
# we make (cash or on behalf of a customer) are credits.

class RawMaterialIn(BaseModel):
    name: str
    unit: Optional[str] = "pcs"          # kg / pcs / litre / m / etc.
    default_rate: float = 0.0            # informational default for purchases
    notes: Optional[str] = ""


class RawMaterialUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    default_rate: Optional[float] = None
    notes: Optional[str] = None


@api_router.get("/raw-materials")
async def list_raw_materials(user=Depends(get_current_user)):
    items = await db.raw_materials.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    return items


@api_router.post("/raw-materials")
async def create_raw_material(body: RawMaterialIn, admin=Depends(require_admin)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "unit": (body.unit or "pcs").strip(),
        "default_rate": round(float(body.default_rate or 0), 2),
        "notes": (body.notes or "").strip(),
        "created_at": now_iso(),
        "created_by": admin["email"],
    }
    await db.raw_materials.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/raw-materials/{rid}")
async def update_raw_material(rid: str, body: RawMaterialUpdate, admin=Depends(require_admin)):
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": admin["email"]}
    if body.name is not None:
        upd["name"] = body.name.strip()
    if body.unit is not None:
        upd["unit"] = body.unit.strip()
    if body.default_rate is not None:
        upd["default_rate"] = round(float(body.default_rate), 2)
    if body.notes is not None:
        upd["notes"] = body.notes.strip()
    res = await db.raw_materials.update_one({"id": rid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return await db.raw_materials.find_one({"id": rid}, {"_id": 0})


@api_router.delete("/raw-materials/{rid}")
async def delete_raw_material(rid: str, admin=Depends(require_admin)):
    res = await db.raw_materials.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return {"ok": True}


class SupplierIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    gst_number: Optional[str] = ""
    contact_person: Optional[str] = ""
    material_category: Optional[str] = ""
    opening_balance: float = 0.0  # +ve = we owe the supplier at start
    notes: Optional[str] = ""


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    gst_number: Optional[str] = None
    contact_person: Optional[str] = None
    material_category: Optional[str] = None
    opening_balance: Optional[float] = None
    notes: Optional[str] = None


class SupplierPurchaseItemIn(BaseModel):
    raw_material_id: Optional[str] = None
    name: str
    unit: Optional[str] = ""
    quantity: float
    rate: float = 0.0  # per-unit price


class SupplierPurchaseIn(BaseModel):
    supplier_id: str
    amount: float
    bill_number: Optional[str] = ""
    purchased_at: Optional[str] = None  # ISO date or datetime; default = now
    material: Optional[str] = ""        # free-text description
    notes: Optional[str] = ""
    items: Optional[List[SupplierPurchaseItemIn]] = None  # optional line-items


class SupplierPaymentIn(BaseModel):
    supplier_id: str
    amount: float
    source: str = "cash"
    reference: Optional[str] = ""
    paid_at: Optional[str] = None
    notes: Optional[str] = ""
    # If set, this payment was made on behalf of a customer; the same amount
    # is credited to that customer's ledger via the linked /payments doc.
    on_behalf_of_customer_id: Optional[str] = None
    customer_payment_id: Optional[str] = None  # back-link to the customers' payment


@api_router.get("/suppliers")
async def list_suppliers(user=Depends(get_current_user)):
    items = await db.suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api_router.post("/suppliers")
async def create_supplier(body: SupplierIn, admin=Depends(require_admin)):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "phone": (body.phone or "").strip(),
        "address": (body.address or "").strip(),
        "city": (body.city or "").strip(),
        "gst_number": (body.gst_number or "").strip(),
        "contact_person": (body.contact_person or "").strip(),
        "material_category": (body.material_category or "").strip(),
        "opening_balance": round(float(body.opening_balance or 0), 2),
        "notes": (body.notes or "").strip(),
        "created_at": now_iso(),
        "created_by": admin["email"],
    }
    await db.suppliers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/suppliers/{sid}")
async def get_supplier(sid: str, user=Depends(get_current_user)):
    s = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return s


@api_router.patch("/suppliers/{sid}")
async def update_supplier(sid: str, body: SupplierUpdate, admin=Depends(require_admin)):
    upd: Dict[str, Any] = {"updated_at": now_iso(), "updated_by": admin["email"]}
    for field in ("name", "phone", "address", "city", "gst_number",
                  "contact_person", "material_category", "notes"):
        val = getattr(body, field, None)
        if val is not None:
            upd[field] = val.strip()
    if body.opening_balance is not None:
        upd["opening_balance"] = round(float(body.opening_balance), 2)
    res = await db.suppliers.update_one({"id": sid}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return await db.suppliers.find_one({"id": sid}, {"_id": 0})


@api_router.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, admin=Depends(require_admin)):
    # Refuse if there are linked purchases or payments
    if await db.supplier_purchases.find_one({"supplier_id": sid}, {"_id": 1}):
        raise HTTPException(status_code=400, detail="Cannot delete supplier with purchase history")
    if await db.supplier_payments.find_one({"supplier_id": sid}, {"_id": 1}):
        raise HTTPException(status_code=400, detail="Cannot delete supplier with payment history")
    res = await db.suppliers.delete_one({"id": sid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"ok": True}


@api_router.post("/supplier-purchases")
async def create_supplier_purchase(body: SupplierPurchaseIn, user=Depends(get_current_user)):
    if not body.supplier_id:
        raise HTTPException(status_code=400, detail="supplier_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    sup = await db.suppliers.find_one({"id": body.supplier_id}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    # Normalize line items if present
    items_out: List[Dict[str, Any]] = []
    if body.items:
        for it in body.items:
            qty = float(it.quantity or 0)
            if qty <= 0 or not (it.name or "").strip():
                continue
            rate = float(it.rate or 0)
            items_out.append({
                "raw_material_id": (it.raw_material_id or None),
                "name": it.name.strip(),
                "unit": (it.unit or "").strip(),
                "quantity": qty,
                "rate": round(rate, 2),
                "line_value": round(rate * qty, 2),
            })
    # If material summary wasn't provided, auto-build one from items
    material = (body.material or "").strip()
    if not material and items_out:
        material = ", ".join(
            f"{it['quantity']} {it['unit']} {it['name']}".strip() for it in items_out
        )
    doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": body.supplier_id,
        "supplier_name": sup.get("name") or "",
        "amount": round(float(body.amount), 2),
        "bill_number": (body.bill_number or "").strip(),
        "material": material,
        "notes": (body.notes or "").strip(),
        "items": items_out,
        "purchased_at": _normalize_payment_dt(body.purchased_at),
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.supplier_purchases.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.delete("/supplier-purchases/{pid}")
async def delete_supplier_purchase(pid: str, admin=Depends(require_admin)):
    res = await db.supplier_purchases.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return {"ok": True}


@api_router.post("/supplier-payments")
async def create_supplier_payment(body: SupplierPaymentIn, user=Depends(get_current_user)):
    if not body.supplier_id:
        raise HTTPException(status_code=400, detail="supplier_id required")
    if body.amount is None or float(body.amount) <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero")
    sup = await db.suppliers.find_one({"id": body.supplier_id}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    customer_id = (body.on_behalf_of_customer_id or "").strip() or None
    customer_name = ""
    customer_payment_id = (body.customer_payment_id or "").strip() or None
    if customer_id:
        cust = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not cust:
            raise HTTPException(status_code=404, detail="Customer not found")
        customer_name = cust.get("name") or ""
    doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": body.supplier_id,
        "supplier_name": sup.get("name") or "",
        "amount": round(float(body.amount), 2),
        "source": _normalize_payment_source(body.source),
        "reference": (body.reference or "").strip(),
        "paid_at": _normalize_payment_dt(body.paid_at),
        "notes": (body.notes or "").strip(),
        "on_behalf_of_customer_id": customer_id,
        "on_behalf_of_customer_name": customer_name,
        "customer_payment_id": customer_payment_id,
        "created_by": user.get("email") or user.get("username") or "",
        "created_at": now_iso(),
    }
    await db.supplier_payments.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.delete("/supplier-payments/{pid}")
async def delete_supplier_payment(pid: str, admin=Depends(require_admin)):
    res = await db.supplier_payments.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier payment not found")
    return {"ok": True}


@api_router.get("/supplier-ledger/{sid}")
async def supplier_ledger(sid: str, user=Depends(get_current_user)):
    """Combined ledger for a single supplier: purchases (Dr) and payments
    (Cr) interleaved by date with a running balance. Balance > 0 means we
    owe the supplier; balance < 0 means the supplier owes us."""
    sup = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    purchases = await db.supplier_purchases.find({"supplier_id": sid}, {"_id": 0}).to_list(5000)
    payments = await db.supplier_payments.find({"supplier_id": sid}, {"_id": 0}).to_list(5000)
    rows: List[Dict[str, Any]] = []
    for p in purchases:
        rows.append({
            "kind": "purchase", "id": p["id"],
            "when": p.get("purchased_at"),
            "debit": float(p.get("amount") or 0), "credit": 0.0,
            "particulars": p.get("material") or "Material purchase",
            "reference": p.get("bill_number") or "",
            "notes": p.get("notes") or "",
            "raw": p,
        })
    for p in payments:
        particulars = "Payment"
        if p.get("on_behalf_of_customer_name"):
            particulars = f"On behalf of {p['on_behalf_of_customer_name']}"
        rows.append({
            "kind": "payment", "id": p["id"],
            "when": p.get("paid_at"),
            "debit": 0.0, "credit": float(p.get("amount") or 0),
            "particulars": particulars,
            "reference": p.get("reference") or "",
            "notes": p.get("notes") or "",
            "raw": p,
        })
    rows.sort(key=lambda r: (r.get("when") or "", r["kind"] == "payment"))
    opening = float(sup.get("opening_balance") or 0)
    bal = opening
    for r in rows:
        bal += r["debit"] - r["credit"]
        r["balance"] = round(bal, 2)
    total_debit = round(sum(r["debit"] for r in rows), 2)
    total_credit = round(sum(r["credit"] for r in rows), 2)
    return {
        "supplier": sup,
        "rows": rows,
        "opening_balance": round(opening, 2),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "closing_balance": round(bal, 2),
    }


# ======================== Health ========================
@api_router.get("/")
async def root():
    return {"message": "Factory Order Management API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    await seed_db()
    # One-time migration: legacy "Delivered" status was removed (Feb 2026).
    # Existing Delivered orders collapse to the new terminal status "Cleared".
    migrated = await db.orders.update_many(
        {"status": "Delivered"}, {"$set": {"status": "Cleared"}}
    )
    if migrated.modified_count:
        logger.info("Migrated %s Delivered → Cleared", migrated.modified_count)

    # One-time backfill: assign sequential numeric slip_no to existing
    # dispatches in chronological order. New dispatches get one via the
    # `next_slip_no()` helper at insert time.
    cursor = db.dispatches.find(
        {"slip_no": {"$exists": False}}, {"_id": 0, "id": 1}
    ).sort("dispatched_at", 1)
    docs = await cursor.to_list(100000)
    if docs:
        # Seed the counter to the current max so we continue from there.
        existing_max = await db.dispatches.find_one(
            {"slip_no": {"$exists": True}}, sort=[("slip_no", -1)], projection={"slip_no": 1, "_id": 0}
        )
        seq = int((existing_max or {}).get("slip_no") or 0)
        for d in docs:
            seq += 1
            await db.dispatches.update_one({"id": d["id"]}, {"$set": {"slip_no": seq}})
        # Sync counter to seq
        await db.counters.update_one(
            {"_id": "dispatch_slip"},
            {"$max": {"seq": seq}},
            upsert=True,
        )
        logger.info("Backfilled slip_no on %s existing dispatches", len(docs))

    # Same backfill for payments → receipt_no
    pcursor = db.payments.find(
        {"receipt_no": {"$exists": False}}, {"_id": 0, "id": 1}
    ).sort("paid_at", 1)
    pdocs = await pcursor.to_list(100000)
    if pdocs:
        existing_max = await db.payments.find_one(
            {"receipt_no": {"$exists": True}}, sort=[("receipt_no", -1)], projection={"receipt_no": 1, "_id": 0}
        )
        seq = int((existing_max or {}).get("receipt_no") or 0)
        for p in pdocs:
            seq += 1
            await db.payments.update_one({"id": p["id"]}, {"$set": {"receipt_no": seq}})
        await db.counters.update_one(
            {"_id": "payment_receipt"},
            {"$max": {"seq": seq}},
            upsert=True,
        )
        logger.info("Backfilled receipt_no on %s existing payments", len(pdocs))


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
