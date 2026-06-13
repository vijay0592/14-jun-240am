"""Seed customers from customer name.xlsx into MongoDB.

Each cell in the Excel sheet is one customer. Format is roughly "NAME-CITY",
"NAME -CITY", "NAME - CITY", or sometimes just "NAME" with no city. We split
on the LAST hyphen to extract city, decode HTML entities, and insert all rows
(duplicates allowed per user instruction).
"""
import asyncio
import html
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

XLSX_PATH = ROOT / "customers.xlsx"
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]


def clean_cell(raw: str) -> tuple[str, str]:
    """Return (name, city) parsed from a single cell. Empty city allowed.

    Strategy: split on the LAST hyphen, but only if it looks like a real
    name/city separator. Heuristics:
      * If hyphen has whitespace on either side, treat as separator.
      * Else split only if the token immediately to the LEFT of the hyphen is
        a "word" (>=3 alphabetic chars). This protects names containing
        hyphens like "A-1", "1-UP", "AAA-10".
    """
    text = html.unescape(raw or "").strip()
    text = re.sub(r"\s+", " ", text)
    if not text:
        return "", ""

    # Find all hyphen positions
    hyphens = [i for i, c in enumerate(text) if c == "-"]
    if not hyphens:
        return text, ""

    last = hyphens[-1]
    left = text[:last]
    right = text[last + 1 :]
    has_space_around = (last > 0 and text[last - 1] == " ") or (
        last + 1 < len(text) and text[last + 1] == " "
    )

    # Token just before hyphen (no spaces), strip leading hyphens/dashes
    left_token_raw = re.split(r"\s+", left.rstrip())[-1] if left.strip() else ""
    left_token = left_token_raw.lstrip("-")

    should_split = False
    if has_space_around:
        should_split = True
    elif re.fullmatch(r"[A-Za-z]{3,}", left_token):
        # tight hyphen like "NIKUNJ-AGRA" or "ZEEPO BIKES-TAMIL NADU"
        should_split = True

    if not should_split:
        return text, ""

    name = left.strip(" -")
    city = right.strip(" -")
    if not city or not re.search(r"[A-Za-z]", city):
        return name or text, ""
    return name, city


def iter_customer_cells(path: Path):
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s or s == "0":
                continue
            yield s


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    parsed = []
    skipped = 0
    for raw in iter_customer_cells(XLSX_PATH):
        name, city = clean_cell(raw)
        if not name:
            skipped += 1
            continue
        parsed.append({"name": name, "city": city, "raw": raw})

    print(f"Parsed {len(parsed)} customers (skipped {skipped} empty cells)")

    # Delete any previously seeded rows from this script (idempotency) so reruns
    # don't double up. We tag inserts with seed_source="xlsx_v1".
    delete_res = await db.customers.delete_many({"seed_source": "xlsx_v1"})
    print(f"Cleared {delete_res.deleted_count} existing xlsx_v1 rows")

    now = datetime.now(timezone.utc).isoformat()
    docs = [
        {
            "id": str(uuid.uuid4()),
            "name": p["name"],
            "phone": "",
            "address": p["city"],
            "preferences": {},
            "created_at": now,
            "seed_source": "xlsx_v1",
        }
        for p in parsed
    ]
    if docs:
        await db.customers.insert_many(docs)
    print(f"Inserted {len(docs)} customers")

    total = await db.customers.count_documents({})
    print(f"Total customers in DB now: {total}")

    # Sample
    print("\nSample (first 8 parsed):")
    for p in parsed[:8]:
        print(f"  name={p['name']!r:40s} city={p['city']!r}")
    print("\nSample (last 5 parsed):")
    for p in parsed[-5:]:
        print(f"  name={p['name']!r:40s} city={p['city']!r}")

    client.close()


if __name__ == "__main__":
    if not XLSX_PATH.exists():
        print(f"Missing file: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)
    asyncio.run(main())
